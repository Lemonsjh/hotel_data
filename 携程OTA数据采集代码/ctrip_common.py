from __future__ import annotations

import os
from datetime import datetime
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

try:
    from ctrip_config import EXTRA_HEADERS
except ImportError:
    EXTRA_HEADERS = {}


ROOT_DIR = Path(os.environ.get("HOTEL_OTA_PROJECT_ROOT") or Path(__file__).resolve().parents[1])
OUTPUT_DIR = Path(os.environ.get("HOTEL_OTA_OUTPUT_DIR") or ROOT_DIR / "OTA数据")
CHANNEL_SOURCE = "携程"


class CtripApiError(RuntimeError):
    pass


class CtripClient:
    def __init__(self, cookie: str, base_url: str = "https://ebooking.ctrip.com"):
        self.cookie = cookie.strip()
        self.base_url = base_url.rstrip("/")
        self.session = requests.Session()
        self.session.headers.update(
            {
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36"
                ),
                "Accept": "application/json, text/plain, */*",
                "Content-Type": "application/json;charset=UTF-8",
                "Origin": "https://ebooking.ctrip.com",
                "Referer": "https://ebooking.ctrip.com/",
            }
        )
        if self.cookie:
            self.session.headers["Cookie"] = self.cookie
        if EXTRA_HEADERS:
            self.session.headers.update(EXTRA_HEADERS)

    def get_json(self, path_or_url: str, params: dict[str, Any] | None = None) -> dict[str, Any]:
        response = self.session.get(self._url(path_or_url), params=params, timeout=30)
        response.raise_for_status()
        return self._decode_json(response)

    def post_json(self, path_or_url: str, payload: dict[str, Any] | None = None) -> dict[str, Any]:
        response = self.session.post(self._url(path_or_url), json=payload or {}, timeout=30)
        response.raise_for_status()
        return self._decode_json(response)

    def _url(self, path_or_url: str) -> str:
        if path_or_url.startswith("http://") or path_or_url.startswith("https://"):
            return path_or_url
        return f"{self.base_url}/{path_or_url.lstrip('/')}"

    @staticmethod
    def _decode_json(response: requests.Response) -> dict[str, Any]:
        try:
            data = response.json()
        except ValueError as exc:
            raise CtripApiError(f"接口没有返回 JSON，HTTP={response.status_code}") from exc
        if not isinstance(data, dict):
            raise CtripApiError("接口 JSON 顶层不是 object")
        return data


def require_config(name: str, value: str) -> str:
    if not value or "TODO" in value:
        raise CtripApiError(f"请先配置 {name}")
    return value


def unwrap_payload(data: dict[str, Any]) -> Any:
    code = data.get("code", data.get("status", data.get("resultCode")))
    success = data.get("success")
    if success is False or code not in (None, 0, "0", 200, "200", "success"):
        message = data.get("message") or data.get("msg") or data.get("errorMsg")
        raise CtripApiError(f"接口返回异常：code={code}, message={message}")
    for key in ("data", "result", "value"):
        if key in data:
            return data[key]
    return data


def parse_number(value: Any) -> Any:
    if value is None or value == "":
        return ""
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip().replace(",", "")
    if text in ("-", "--"):
        return "-"
    if text.endswith("%"):
        return parse_percent(text)
    try:
        number = float(text)
        return int(number) if number.is_integer() else number
    except ValueError:
        return value


def parse_percent(value: Any) -> Any:
    if value is None or value == "":
        return ""
    if isinstance(value, (int, float)):
        return value if abs(value) <= 1 else value / 100
    text = str(value).strip().replace(",", "")
    if text in ("-", "--"):
        return "-"
    if text.endswith("%"):
        text = text[:-1]
    try:
        return float(text) / 100
    except ValueError:
        return value


def join_unique(values: list[Any], sep: str = "、") -> str:
    result: list[str] = []
    for value in values:
        text = str(value).strip() if value is not None else ""
        if text and text not in result:
            result.append(text)
    return sep.join(result)


def write_single_sheet(
    output_path: Path,
    sheet_name: str,
    headers: list[str],
    rows: list[list[Any]],
    widths: list[int] | None = None,
    percent_columns: set[int] | None = None,
    datetime_columns: set[int] | None = None,
    date_columns: set[int] | None = None,
) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append(row)
    for idx, width in enumerate(widths or [18] * len(headers), 1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if percent_columns and cell.column in percent_columns:
                cell.number_format = "0.00%"
            if datetime_columns and cell.column in datetime_columns:
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
            if date_columns and cell.column in date_columns:
                cell.number_format = "yyyy-mm-dd"
    wb.save(output_path)
    return output_path


def now_snapshot() -> datetime:
    return datetime.now()
