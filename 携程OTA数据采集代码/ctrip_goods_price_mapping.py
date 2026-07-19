from __future__ import annotations

import argparse
import json
import os
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from ctrip_config import COOKIE, EXTRA_HEADERS, PLATFORM_SCOPE, USER_AGENT

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from ota_mysql_writer import OUTPUT_DIR, sync_table


TABLE_NAME = "ctrip_ota_goods_price_mapping"
HOTEL_ID = os.environ.get("HOTEL_ID", "").strip()
CHANNEL_SOURCE = "\u643a\u7a0b"

DEFAULT_QUERY_URL = (
    "https://ebooking.ctrip.com/restapi/soa2/30535/getRCRoomProductList"
    "?_fxpcqlniredt=09031067317598779101"
    "&x-traceID=09031067317598779101-1782364705290-2636122"
)
QUERY_URL = os.environ.get("CTRIP_GOODS_QUERY_URL", DEFAULT_QUERY_URL).strip()
INVENTORY_URL = "https://ebooking.ctrip.com/ebkovsroom/api/inventory/getRoomInventoryInfo"

HEADERS = [
    "snapshot_time",
    "channel_source",
    "ota_hotel_id",
    "ota_room_type_id",
    "room_type_name",
    "business_date",
    "ota_product_id",
    "ota_product_name",
    "product_cipher",
    "price_editable_flag",
    "is_hour_room",
    "ota_sale_price",
    "commission_rate",
]


class CtripGoodsClient:
    def __init__(self, cookie: str, query_url: str = QUERY_URL):
        self.query_url = query_url
        self.session = requests.Session()
        self.session.headers.update(
            {
                "User-Agent": USER_AGENT,
                "Accept": "application/json, text/plain, */*",
                "Content-Type": "application/json;charset=UTF-8",
                "Origin": "https://ebooking.ctrip.com",
                "Referer": "https://ebooking.ctrip.com/ebkovsroom/inventory/roompricemanagement?microJump=true",
            }
        )
        if cookie.strip():
            self.session.headers["Cookie"] = cookie.strip()
        if EXTRA_HEADERS:
            self.session.headers.update(EXTRA_HEADERS)

    def query_goods(self) -> dict[str, Any]:
        response = self.session.post(self.query_url, json={}, timeout=30)
        response.raise_for_status()
        payload = response.json()
        if not isinstance(payload, dict):
            raise RuntimeError("getRCRoomProductList response is not object")
        status = payload.get("ResponseStatus") or {}
        if status and status.get("Ack") not in (None, "Success"):
            raise RuntimeError({"ResponseStatus": status})
        return payload.get("data") if isinstance(payload.get("data"), dict) else payload

    def query_inventory(self, products: list[dict[str, Any]], start_date: str, end_date: str) -> dict[str, Any]:
        payload = {
            "startDate": start_date,
            "endDate": end_date,
            "showRoomPrice": True,
            "showRoomInventory": True,
            "showLadderPolicy": True,
            "isPreTaxPrice": False,
            "saleChannel": 0,
            "hotelRoomInfoDtoList": [
                {"hotelID": int(item["sub_hotel_id"]), "roomTypeID": int(item["ota_product_id"])}
                for item in products
                if item.get("sub_hotel_id") and item.get("ota_product_id")
            ],
        }
        response = self.session.post(INVENTORY_URL, json=payload, timeout=30)
        response.raise_for_status()
        payload = response.json()
        if payload.get("code") != 200:
            raise RuntimeError({"code": payload.get("code"), "message": payload.get("message")})
        return payload.get("data") if isinstance(payload.get("data"), dict) else payload


def clean_name(value: Any) -> str:
    return "" if value is None else str(value).strip()


ROOM_NAME_SUFFIX_RE = re.compile(r"\s*(?:\[[^\[\]]*\]|【[^【】]*】|\([^()]*\)|（[^（）]*）)\s*$")


def base_room_type_name(value: Any) -> str:
    name = clean_name(value)
    while name:
        cleaned = ROOM_NAME_SUFFIX_RE.sub("", name).strip()
        if cleaned == name:
            break
        name = cleaned
    return name


def ota_product_name(product: dict[str, Any]) -> str:
    name = clean_name(product.get("productDisplayName"))
    suffix = f"({product.get('productId')})"
    return name.removesuffix(suffix).strip()


def normalize_products(payload: dict[str, Any]) -> list[dict[str, Any]]:
    cipher_map = payload.get("cipher") or {}
    room_map = payload.get("basicRoomTypeMap") or {}
    product_map = payload.get("roomProducts") or {}
    products: list[dict[str, Any]] = []
    for room_id in payload.get("basicRoomIds") or room_map.keys():
        room = room_map.get(str(room_id)) or {}
        if not isinstance(room, dict):
            continue
        product_ids = room.get("productIds") or []
        for ota_product_id in product_ids:
            product = product_map.get(str(ota_product_id)) or {}
            if not isinstance(product, dict):
                continue
            products.append(
                {
                    "hotel_id": room.get("masterHotelId") or product.get("subHotelId"),
                    "sub_hotel_id": product.get("subHotelId"),
                    "ota_room_type_id": room.get("basicRoomId") or product.get("masterBasicRoomId"),
                    "room_type_name": base_room_type_name(room.get("basicRoomName")),
                    "ota_product_id": product.get("productId") or ota_product_id,
                    "ota_product_name": ota_product_name(product),
                    "product_cipher": cipher_map.get(str(ota_product_id), ""),
                    "price_editable_flag": product.get("editPrice"),
                    "is_hour_room": product.get("hourRoom"),
                }
            )
    return products


def price_index(inventory_payload: dict[str, Any]) -> dict[tuple[str, str], dict[str, Any]]:
    infos = ((inventory_payload.get("roomPriceResult") or {}).get("roomPriceInfo") or [])
    result: dict[tuple[str, str], dict[str, Any]] = {}
    for item in infos:
        if isinstance(item, dict):
            result[(str(item.get("roomTypeID")), str(item.get("effectDate")))] = item
    return result


def normalize_rows(
    products: list[dict[str, Any]],
    captured_at: datetime,
    inventory_payload: dict[str, Any] | None = None,
) -> list[list[Any]]:
    prices = price_index(inventory_payload or {})
    dates = sorted({key[1] for key in prices}) or [""]
    rows: list[list[Any]] = []
    for product in products:
        for business_date in dates:
            price = prices.get((str(product.get("ota_product_id")), business_date), {})
            rows.append(
                [
                    captured_at,
                    CHANNEL_SOURCE,
                    product.get("hotel_id"),
                    product.get("ota_room_type_id"),
                    product.get("room_type_name"),
                    business_date,
                    product.get("ota_product_id"),
                    product.get("ota_product_name"),
                    product.get("product_cipher"),
                    product.get("price_editable_flag"),
                    product.get("is_hour_room"),
                    price.get("originalPrice"),
                    price.get("commissionRatePercent") or price.get("commissionRate"),
                ]
            )
    return rows


def json_safe(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return value


def write_json(path: Path, headers: list[str], rows: list[list[Any]]) -> Path:
    json_path = path.with_suffix(".json")
    payload = {
        "table_name": path.stem,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "row_count": len(rows),
        "rows": [
            {header: json_safe(row[index]) if index < len(row) else "" for index, header in enumerate(headers)}
            for row in rows
        ],
    }
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return json_path


def save_excel(rows: list[list[Any]]) -> Path:
    headers = [*HEADERS, "platform_scope", "hotel_id"]
    rows = [list(row) + [PLATFORM_SCOPE, HOTEL_ID] for row in rows]
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUTPUT_DIR / f"{TABLE_NAME}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = TABLE_NAME
    ws.append(headers)
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append(row)
        ws.cell(ws.max_row, 1).number_format = "yyyy-mm-dd hh:mm:ss"
    widths = [20, 12, 14, 16, 38, 16, 16, 50, 72, 12, 14, 18, 22, 16]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=index).column_letter].width = width
    wb.save(out_path)
    write_json(out_path, headers, rows)
    sync_table(TABLE_NAME, headers, rows)
    return out_path


def load_sample_file(path: Path) -> dict[str, Any]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise RuntimeError("sample file JSON is not object")
    return payload.get("data") if isinstance(payload.get("data"), dict) else payload


def collect(
    query_url: str = QUERY_URL,
    sample_file: Path | None = None,
    inventory_sample_file: Path | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
) -> list[list[Any]]:
    captured_at = datetime.now()
    client = CtripGoodsClient(COOKIE, query_url)
    payload = load_sample_file(sample_file) if sample_file else client.query_goods()
    products = normalize_products(payload)
    if inventory_sample_file:
        inventory_payload = load_sample_file(inventory_sample_file)
    else:
        start = start_date or captured_at.strftime("%Y-%m-%d")
        end = end_date or start
        inventory_payload = client.query_inventory(products, start, end)
    return normalize_rows(products, captured_at, inventory_payload)


def main() -> None:
    parser = argparse.ArgumentParser(description="Ctrip room product price mapping crawler.")
    parser.add_argument("--url", default=QUERY_URL, help="getRCRoomProductList URL copied from Network.")
    parser.add_argument("--sample-file", type=Path, help="Parse saved response JSON; no request.")
    parser.add_argument("--inventory-sample-file", type=Path, help="Parse saved getRoomInventoryInfo JSON.")
    parser.add_argument("--start-date", help="Inventory start date, YYYY-MM-DD.")
    parser.add_argument("--end-date", help="Inventory end date, YYYY-MM-DD.")
    args = parser.parse_args()
    rows = collect(args.url, args.sample_file, args.inventory_sample_file, args.start_date, args.end_date)
    out_path = save_excel(rows)
    print(f"room product rows={len(rows)} output={out_path}")


if __name__ == "__main__":
    main()
