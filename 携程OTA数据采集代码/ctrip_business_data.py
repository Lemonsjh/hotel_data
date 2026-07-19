from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from ctrip_config import COOKIE, DEFAULT_HOTEL_NAME, EXTRA_HEADERS, PLATFORM_SCOPE, USER_AGENT

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from ota_mysql_writer import OUTPUT_DIR, sync_ctrip_metric_history


ENDPOINTS = {
    "order_loss": "/restapi/soa2/24588/getTripartiteOrderLoss",
    "visitor_title": "/datacenter/api/dataCenter/current/fetchVisitorTitleV2",
    "management_data": "/restapi/soa2/24588/getManagementData",
    "flow_data": "/restapi/soa2/24588/getFlowData",
}
DEFAULT_OUTPUT = OUTPUT_DIR / "ctrip_ota_business_metrics.xlsx"
COMPETITION_CAPTURE = OUTPUT_DIR / "ctrip_competition_profile_dom_capture.json"
COMPETITION_REALTIME_CAPTURE = OUTPUT_DIR / "ctrip_competition_profile_realtime_dom_capture.json"
SHEET_NAME = "采集明细"
HOTEL_ID = os.environ.get("HOTEL_ID", "").strip()

HEADERS = [
    "snapshot_time",
    "business_date",
    "metric_code",
    "hotel_name",
    "metric_group",
    "metric_name",
    "metric_value",
    "metric_unit",
    "compare_label",
    "compare_value",
    "competitor_rank",
    "peer_average",
]

METRIC_LABELS = {
    "booking_order_count": ("\u7ecf\u8425\u5bf9\u6bd4", "\u9884\u8ba2\u8ba2\u5355\u91cf"),
    "booking_sales_amount": ("\u7ecf\u8425\u5bf9\u6bd4", "\u9884\u8ba2\u9500\u552e\u989d"),
    "inhouse_room_night": ("\u7ecf\u8425\u5bf9\u6bd4", "\u5728\u5e97\u95f4\u591c"),
    "occupancy_rate": ("\u7ecf\u8425\u5bf9\u6bd4", "\u51fa\u79df\u7387"),
    "ctrip_app_visitor_count": ("\u7ecf\u8425\u5bf9\u6bd4", "\u643a\u7a0bAPP\u8bbf\u5ba2"),
    "ctrip_app_conversion_rate": ("\u7ecf\u8425\u5bf9\u6bd4", "\u643a\u7a0bAPP\u8f6c\u5316\u7387"),
    "realtime_booking_sales_amount": ("\u5b9e\u65f6\u7ecf\u8425", "\u5b9e\u65f6\u9884\u8ba2\u9500\u552e\u989d"),
    "realtime_occupancy_rate": ("\u5b9e\u65f6\u7ecf\u8425", "\u5b9e\u65f6\u51fa\u79df\u7387"),
    "realtime_booking_order_count": ("实时经营", "实时预订订单"),
    "realtime_inhouse_room_night": ("实时经营", "实时在店间夜"),
    "realtime_ctrip_app_visitor_count": ("\u5b9e\u65f6\u7ecf\u8425", "\u5b9e\u65f6\u643a\u7a0bAPP\u8bbf\u5ba2"),
    "realtime_ctrip_app_conversion_rate": ("\u5b9e\u65f6\u7ecf\u8425", "\u5b9e\u65f6\u643a\u7a0bAPP\u8f6c\u5316\u7387"),
    "list_page_exposure_count": ("流量漏斗", "列表页曝光量"),
    "detail_page_visitor_count": ("流量漏斗", "详情页访客量"),
    "order_submit_count": ("流量漏斗", "订单量"),
    "exposure_conversion_rate": ("流量漏斗", "曝光转化率"),
    "order_conversion_rate": ("流量漏斗", "下单转化率"),
    "lost_order_count": ("流失诊断", "流失订单量"),
    "lost_room_night_count": ("流失诊断", "流失间夜量"),
    "lost_order_amount": ("流失诊断", "流失订单金额"),
}


class CtripApiError(RuntimeError):
    pass


class CtripClient:
    def __init__(self, cookie: str, base_url: str = "https://ebooking.ctrip.com"):
        self.base_url = base_url.rstrip("/")
        self.session = requests.Session()
        self.session.headers.update(
            {
                "User-Agent": USER_AGENT,
                "Accept": "application/json, text/plain, */*",
                "Content-Type": "application/json;charset=UTF-8",
                "Origin": "https://ebooking.ctrip.com",
                "Referer": "https://ebooking.ctrip.com/",
            }
        )
        if cookie.strip():
            self.session.headers["Cookie"] = cookie.strip()
        if EXTRA_HEADERS:
            self.session.headers.update(EXTRA_HEADERS)

    def post_json(
        self,
        path_or_url: str,
        payload: dict[str, Any] | None = None,
        content_type: str | None = None,
    ) -> Any:
        if content_type:
            response = self.session.post(
                self._url(path_or_url),
                data=json.dumps(payload or {}),
                headers={"Content-Type": content_type},
                timeout=30,
            )
        else:
            response = self.session.post(self._url(path_or_url), json=payload or {}, timeout=30)
        response.raise_for_status()
        try:
            data = response.json()
        except ValueError as exc:
            raise CtripApiError(f"接口没有返回 JSON，HTTP={response.status_code}") from exc
        return data

    def _url(self, path_or_url: str) -> str:
        if path_or_url.startswith("http://") or path_or_url.startswith("https://"):
            return path_or_url
        return f"{self.base_url}/{path_or_url.lstrip('/')}"


def require_config(name: str, value: str) -> str:
    if not value or "TODO" in value:
        raise CtripApiError(f"请先配置 {name}")
    return value


def unwrap_payload(data: dict[str, Any]) -> Any:
    code = data.get("code", data.get("status", data.get("resultCode")))
    success = data.get("success")
    if success is False or code not in (None, 0, "0", 200, "200", "success"):
        message = data.get("message") or data.get("msg") or data.get("errorMsg")
        raise CtripApiError(f"接口返回异常：code={code}, message={message}")
    for key in ("data", "result", "value"):
        if key in data:
            return data[key]
    return data


def parse_number(value: Any) -> Any:
    if value is None or value == "":
        return ""
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip().replace(",", "")
    if text in ("-", "--"):
        return "-"
    if text.endswith("%"):
        text = text[:-1]
        try:
            return float(text) / 100
        except ValueError:
            return value
    try:
        number = float(text)
        return int(number) if number.is_integer() else number
    except ValueError:
        return value


def normalize_metric_value(value: Any, unit: str) -> Any:
    if unit == "%" and isinstance(value, str) and value.strip().endswith("%"):
        return parse_number(value.strip()[:-1])
    return parse_number(value)


def round_number(value: Any, digits: int = 2) -> Any:
    parsed = parse_number(value)
    if isinstance(parsed, float):
        parsed = round(parsed, digits)
        return int(parsed) if parsed.is_integer() else parsed
    return parsed


def percent_points(value: Any) -> Any:
    parsed = parse_number(value)
    if isinstance(parsed, (int, float)):
        return round_number(parsed * 100)
    return parsed


def write_single_sheet(
    output_path: Path,
    sheet_name: str,
    headers: list[str],
    rows: list[list[Any]],
    widths: list[int],
    datetime_columns: set[int] | None = None,
    date_columns: set[int] | None = None,
) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append(row)
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if datetime_columns and cell.column in datetime_columns:
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
            if date_columns and cell.column in date_columns:
                cell.number_format = "yyyy-mm-dd"
    wb.save(output_path)
    return output_path


def json_safe(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if hasattr(value, "isoformat") and not isinstance(value, (str, int, float, bool)):
        return value.isoformat()
    return value


def write_standard_json(output_path: Path, headers: list[str], rows: list[list[Any]]) -> Path:
    json_path = output_path.with_suffix(".json")
    payload = {
        "table_name": output_path.stem,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "row_count": len(rows),
        "rows": [
            {header: json_safe(row[index]) if index < len(row) else "" for index, header in enumerate(headers)}
            for row in rows
        ],
    }
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return json_path


class CtripBusinessClient(CtripClient):
    def optional_post_json(
        self,
        path_or_url: str,
        payload: dict[str, Any] | None = None,
        content_type: str | None = None,
    ) -> Any:
        try:
            return self.post_json(path_or_url, payload, content_type)
        except Exception as exc:
            print(f"WARN optional endpoint skipped: {path_or_url} ({exc})")
            return {}

    def query_all(self) -> dict[str, Any]:
        yesterday = (datetime.now().date() - timedelta(days=1)).strftime("%Y-%m-%d")
        data = {
            name: self.post_json(path, {})
            for name, path in ENDPOINTS.items()
            if name not in {"order_loss", "flow_data", "management_data", "management_data_realtime"}
        }
        data["management_data"] = self.post_json(
            ENDPOINTS["management_data"],
            {
                "dateType": 2,
                "beginDate": yesterday,
                "endDate": yesterday,
                "cipher": {},
                "header": {"platform": "WEB"},
            },
        )
        data["management_data_realtime"] = self.post_json(
            ENDPOINTS["management_data"],
            {
                "dateType": 1,
                "beginDate": "",
                "endDate": "",
                "cipher": {},
                "header": {"platform": "WEB"},
            },
        )
        data["order_loss"] = self.post_json(
            ENDPOINTS["order_loss"],
            {
                "ota": "ctrip",
                "dateType": 2,
                "beginDate": yesterday,
                "endDate": yesterday,
                "pageNo": 1,
                "pageSize": 10,
                "sortKey": 0,
                "desc": 2,
                "cipher": {},
                "header": {"platform": "WEB"},
            },
        )
        data["flow_data"] = self.post_json(
            ENDPOINTS["flow_data"],
            {
                "dateType": 2,
                "beginDate": yesterday,
                "endDate": yesterday,
                "ota": "ctrip",
                "cipher": {},
                "header": {"platform": "WEB"},
            },
        )
        return data


def payload_data(response: Any) -> Any:
    if isinstance(response, dict) and "data" in response:
        return response.get("data")
    return response


def row(
    captured_at: datetime,
    business_date: Any,
    hotel_name: str,
    metric_code: str,
    metric_value: Any,
    metric_unit: str = "",
    compare_label: str = "",
    compare_value: Any = "",
    competitor_rank: Any = "",
    peer_average: Any = "",
) -> list[Any]:
    metric_group, metric_name = METRIC_LABELS.get(metric_code, ("其他", metric_code))
    return [
        captured_at,
        business_date,
        metric_code,
        hotel_name,
        metric_group,
        metric_name,
        normalize_metric_value(metric_value, metric_unit),
        metric_unit,
        compare_label,
        compare_value,
        competitor_rank,
        normalize_metric_value(peer_average, metric_unit) if metric_unit == "%" else peer_average,
    ]


COMPETITION_TITLE_MAP = {
    "\u643a\u7a0bAPP\u8bbf\u5ba2": ("ctrip_app_visitor_count", "person"),
    "\u643a\u7a0bAPP\u8f6c\u5316\u7387": ("ctrip_app_conversion_rate", "%"),
}


def load_competition_capture(path: Path = COMPETITION_CAPTURE) -> dict[str, Any]:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}


def load_competition_captures() -> list[dict[str, Any]]:
    captures = []
    for path in (COMPETITION_REALTIME_CAPTURE, COMPETITION_CAPTURE):
        capture = load_competition_capture(path)
        if capture:
            captures.append(capture)
    return captures


def normalize_compare_text(text: Any) -> tuple[str, Any]:
    if not text:
        return "", ""
    value = str(text).strip()
    for label in ("\u8f83\u4e0a\u5468\u540c\u671f", "\u8f83\u53bb\u5e74\u540c\u671f", "\u8f83\u6628\u65e5"):
        if value.startswith(label):
            return label, value[len(label) :].strip()
    return "", value


def competition_profile_rows(data: dict[str, Any], captured_at: datetime, business_date: Any, hotel_name: str) -> list[list[Any]]:
    metrics = data.get("metrics") if isinstance(data, dict) else []
    if not isinstance(metrics, list):
        return []
    period_type = data.get("statsPeriodType") or "yesterday"
    period_business_date = captured_at.date() if period_type == "realtime" else business_date
    rows: list[list[Any]] = []
    for item in metrics:
        if not isinstance(item, dict):
            continue
        mapped = COMPETITION_TITLE_MAP.get(str(item.get("title", "")).strip())
        if not mapped:
            continue
        metric_name, metric_unit = mapped
        compare_label, compare_value = normalize_compare_text(item.get("hotelCompare"))
        rows.append(
            row(
                captured_at,
                period_business_date,
                hotel_name,
                metric_name,
                item.get("hotelValue"),
                metric_unit,
                compare_label,
                compare_value,
                item.get("competitorRank", ""),
                parse_number(item.get("peerAverage", "")),
            )
        )
    return rows


def extract_hotel_name_from_api(data: dict[str, Any]) -> str | None:
    """尝试从各API接口响应中提取酒店名。"""
    # 1. 尝试 visitor_title 接口
    visitor_title = payload_data(data.get("visitor_title"))
    if isinstance(visitor_title, dict):
        for key in ("hotelName", "hotel_name", "name", "title", "hotelTitle"):
            value = visitor_title.get(key)
            if value:
                return str(value).strip()
    # 2. 尝试标题接口中的酒店名称
    resp = payload_data(data.get("visitor_title"))
    if isinstance(resp, dict):
        for key in ("hotelName", "hotel_name", "name", "hotelTitle", "hotelIdName"):
            value = resp.get(key)
            if value:
                return str(value).strip()
    return None


def extract_hotel_name(payload: dict[str, Any], override_name: str | None = None) -> str:
    """提取酒店名，优先级：外部传入 > API响应 > 竞对数据 > 默认值。"""
    if override_name:
        return override_name.strip()

    # 从顶层 payload 取
    for key in ("hotelName", "hotel_name"):
        value = payload.get(key)
        if value:
            return str(value).strip()

    # 从各API接口尝试提取
    api_name = extract_hotel_name_from_api(payload)
    if api_name:
        return api_name

    # 从竞对分析数据中取
    profiles = payload.get("competition_profiles") or []
    if payload.get("competition_profile"):
        profiles = [payload["competition_profile"], *profiles]
    for profile in profiles:
        if isinstance(profile, dict):
            for key in ("hotelName", "hotel_name"):
                value = profile.get(key)
                if value:
                    return str(value).strip()

    return DEFAULT_HOTEL_NAME


def business_diagnosis_rows(data: dict[str, Any], captured_at: datetime, business_date: Any, hotel_name: str) -> list[list[Any]]:
    response = payload_data(data.get("order_loss")) or {}
    loss = response.get("orderLossdata") or {}
    if not loss:
        return []
    return [
        row(captured_at, business_date, hotel_name, "lost_order_count", loss.get("order"), "order", "较前日", round_number(loss.get("orderYoy"))),
        row(captured_at, business_date, hotel_name, "lost_room_night_count", loss.get("quantity"), "room_night", "较前日", round_number(loss.get("quantityYoy"))),
        row(captured_at, business_date, hotel_name, "lost_order_amount", loss.get("amount"), "CNY", "较前日", round_number(loss.get("amountYoy"))),
    ]


MANAGEMENT_DATA_METRICS = {
    0: ("booking_order_count", "order"),
    1: ("booking_sales_amount", "CNY"),
    2: ("inhouse_room_night", "room_night"),
    3: ("occupancy_rate", "%"),
    4: ("ctrip_app_visitor_count", "person"),
    5: ("ctrip_app_conversion_rate", "%"),
}

REALTIME_MANAGEMENT_DATA_METRICS = {
    0: ("realtime_booking_order_count", "order"),
    1: ("realtime_booking_sales_amount", "CNY"),
    2: ("realtime_inhouse_room_night", "room_night"),
    3: ("realtime_occupancy_rate", "%"),
    4: ("realtime_ctrip_app_visitor_count", "person"),
    5: ("realtime_ctrip_app_conversion_rate", "%"),
}


def management_data_rows(
    data: dict[str, Any],
    captured_at: datetime,
    business_date: Any,
    hotel_name: str,
    realtime: bool = False,
) -> list[list[Any]]:
    response_key = "management_data_realtime" if realtime else "management_data"
    response = payload_data(data.get(response_key)) or {}
    items = response.get("dataList") or []
    metrics = REALTIME_MANAGEMENT_DATA_METRICS if realtime else MANAGEMENT_DATA_METRICS
    compare_label = "\u8f83\u6628\u65e5\u540c\u671f" if realtime else "\u8f83\u4e0a\u5468\u540c\u671f"
    rows: list[list[Any]] = []
    for item in items:
        if not isinstance(item, dict) or item.get("indexType") not in metrics:
            continue
        metric_code, metric_unit = metrics[item["indexType"]]
        is_percentage = metric_unit == "%"
        value = lambda key: percent_points(item.get(key)) if is_percentage else round_number(item.get(key))
        rows.append(
            row(
                captured_at,
                business_date,
                hotel_name,
                metric_code,
                value("val"),
                metric_unit,
                compare_label,
                value("lastVal"),
                item.get("rankComp", ""),
                value("avgComp"),
            )
        )
    return rows


FLOW_DATA_METRICS = {
    6: ("list_page_exposure_count", "count"),
    7: ("detail_page_visitor_count", "person"),
    8: ("order_submit_count", "order"),
    9: ("exposure_conversion_rate", "%"),
    10: ("order_conversion_rate", "%"),
}


def flow_data_rows(data: dict[str, Any], captured_at: datetime, business_date: Any, hotel_name: str) -> list[list[Any]]:
    response = payload_data(data.get("flow_data")) or {}
    items = response.get("dataList") or []
    rows: list[list[Any]] = []
    for item in items:
        if not isinstance(item, dict) or item.get("indexType") not in FLOW_DATA_METRICS:
            continue
        metric_code, metric_unit = FLOW_DATA_METRICS[item["indexType"]]
        is_percentage = metric_unit == "%"
        rows.append(
            row(
                captured_at,
                business_date,
                hotel_name,
                metric_code,
                percent_points(item.get("val")) if is_percentage else item.get("val"),
                metric_unit,
                "较前日",
                percent_points(item.get("lastVal")) if is_percentage else item.get("lastVal"),
                item.get("rankComp", ""),
                percent_points(item.get("avgComp")) if is_percentage else item.get("avgComp"),
            )
        )
    return rows


def normalize_rows(payload: dict[str, Any], captured_at: datetime, hotel_name: str | None = None) -> list[list[Any]]:
    business_date = captured_at.date() - timedelta(days=1)
    hotel_name = extract_hotel_name(payload, override_name=hotel_name)
    profiles = payload.get("competition_profiles") or []
    if not profiles and payload.get("competition_profile"):
        profiles = [payload["competition_profile"]]
    rows: list[list[Any]] = []
    rows.extend(management_data_rows(payload, captured_at, business_date, hotel_name))
    rows.extend(management_data_rows(payload, captured_at, captured_at.date(), hotel_name, realtime=True))
    for profile in profiles:
        rows.extend(competition_profile_rows(profile, captured_at, business_date, hotel_name))
    rows.extend(flow_data_rows(payload, captured_at, business_date, hotel_name))
    rows.extend(business_diagnosis_rows(payload, captured_at, business_date, hotel_name))
    return rows


def deduplicate_rows(rows: list[list[Any]]) -> list[list[Any]]:
    """同一营业日和指标按组装顺序保留最后一条。"""
    unique: dict[tuple[Any, Any], list[Any]] = {}
    for item in rows:
        unique[(item[1], item[2])] = item
    return list(unique.values())


def sample_payload() -> dict[str, Any]:
    return {
        "hotelName": "示例酒店",
        "metrics": [
            {"name": "销售额", "value": "1234.00", "unit": "元", "compareLabel": "较昨日", "compareValue": "+8.2%", "rank": "3/20", "peerAvg": "980.00"},
            {"name": "订单数", "value": "12", "unit": "单", "compareLabel": "较昨日", "compareValue": "-1", "rank": "5/20", "peerAvg": "10"},
            {"name": "支付转化率", "value": "6.50%", "unit": "%", "compareLabel": "较昨日", "compareValue": "+0.8pp", "rank": "6/20", "peerAvg": "5.80%"},
        ],
    }


def save_rows(rows: list[list[Any]], output: Path = DEFAULT_OUTPUT) -> Path:
    if not HOTEL_ID:
        raise CtripApiError("HOTEL_ID 未配置，拒绝写入携程经营历史")
    rows = deduplicate_rows(rows)
    if not rows:
        raise CtripApiError("携程经营接口未生成有效指标，保留数据库原有数据")
    headers = [*HEADERS[:4], "hotel_id", "platform_scope", *HEADERS[4:]]
    rows = [list(item[:4]) + [HOTEL_ID, PLATFORM_SCOPE] + list(item[4:]) for item in rows]
    output_path = write_single_sheet(
        output,
        SHEET_NAME,
        headers,
        rows,
        widths=[20, 14, 32, 24, 16, 16, 22, 14, 8, 14, 14, 12, 14],
        datetime_columns={1},
        date_columns={2},
    )
    write_standard_json(output_path, headers, rows)
    sync_ctrip_metric_history(headers, rows)
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(description="携程经营概览采集框架")
    parser.add_argument("--self-test", action="store_true", help="使用样例 JSON 测试解析和 Excel 输出")
    parser.add_argument("--output", default=str(DEFAULT_OUTPUT))
    parser.add_argument("--cookie", default=COOKIE)
    parser.add_argument("--hotel-name", default="", help="指定酒店名称，覆盖默认值和API返回值")
    args = parser.parse_args()

    captured_at = datetime.now()
    if args.self_test:
        payload = sample_payload()
    else:
        payload = CtripBusinessClient(args.cookie).query_all()
    rows = normalize_rows(payload, captured_at, hotel_name=args.hotel_name or None)
    output = save_rows(rows, Path(args.output))
    print(f"OK 携程经营指标行数={len(rows)} 输出={output}")


if __name__ == "__main__":
    main()
