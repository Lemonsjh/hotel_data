from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from ctrip_config import COOKIE, EXTRA_HEADERS, USER_AGENT

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from ota_mysql_writer import OUTPUT_DIR, sync_table


CHANNEL_SOURCE = "\u643a\u7a0b"
HOTEL_ID = os.environ.get("HOTEL_ID", "").strip()
GET_PRO_BATCH_URL = os.environ.get(
    "CTRIP_GET_PRO_BATCH_URL",
    "https://ebooking.ctrip.com/restapi/soa2/24267/getProBatch"
    "?_fxpcqlniredt=09031067317598779101"
    "&x-traceID=09031067317598779101-1782387657114-1720575",
).strip()

PROMOTION_OUTPUT = OUTPUT_DIR / "ctrip_ota_promotion_activity.xlsx"
DETAIL_OUTPUT = OUTPUT_DIR / "ctrip_ota_activity_product_detail.xlsx"
SUMMARY_SHEET = "\u4fc3\u9500\u6d3b\u52a8"
DETAIL_SHEET = "\u6d3b\u52a8\u4ea7\u54c1\u660e\u7ec6"

SUMMARY_HEADERS = [
    "snapshot_time",
    "channel_source",
    "activity_source_type",
    "activity_name",
    "activity_status",
    "activity_time_range",
    "activity_rule_labels",
    "activity_room_type_summary",
]

DETAIL_HEADERS = [
    "snapshot_time",
    "channel_source",
    "activity_source_type",
    "activity_name",
    "ota_room_type_id",
    "room_type_name",
    "remaining_inventory",
]

STATUS_MAP = {
    "EFFECTIVE": "\u5df2\u751f\u6548",
    "AUDIT": "\u5ba1\u6838\u4e2d",
    "EXPIRED": "\u5df2\u8fc7\u671f",
    "PAUSE": "\u5df2\u6682\u505c",
}


class CtripApiError(RuntimeError):
    pass


class CtripPromotionClient:
    def __init__(self, cookie: str):
        self.session = requests.Session()
        self.session.headers.update(
            {
                "User-Agent": USER_AGENT,
                "Accept": "application/json, text/plain, */*",
                "Content-Type": "application/json;charset=UTF-8",
                "Origin": "https://ebooking.ctrip.com",
                "Referer": "https://ebooking.ctrip.com/promotion/promotionCenter?microJump=true",
            }
        )
        if cookie.strip():
            self.session.headers["Cookie"] = cookie.strip()
        if EXTRA_HEADERS:
            self.session.headers.update(EXTRA_HEADERS)

    def query_promotion_batch(self, url: str = GET_PRO_BATCH_URL) -> dict[str, Any]:
        response = self.session.post(url, json=build_payload(), timeout=30)
        response.raise_for_status()
        try:
            data = response.json()
        except ValueError as exc:
            raise CtripApiError(f"\u63a5\u53e3\u6ca1\u6709\u8fd4\u56de JSON\uff0cHTTP={response.status_code}") from exc
        if not isinstance(data, dict):
            raise CtripApiError("\u63a5\u53e3 JSON \u9876\u5c42\u4e0d\u662f object")
        status = data.get("ResponseStatus") or {}
        if isinstance(status, dict) and status.get("Ack") not in (None, "Success"):
            raise CtripApiError(f"getProBatch \u8fd4\u56de\u5f02\u5e38\uff1a{status.get('Ack')}")
        res_status = data.get("resStatus") or {}
        code = res_status.get("rcode") if isinstance(res_status, dict) else None
        if code not in (None, 0, "0", 200, "200"):
            raise CtripApiError(f"getProBatch \u8fd4\u56de\u5f02\u5e38\uff1arcode={code}")
        return data


def build_payload() -> dict[str, Any]:
    return {
        "reqHead": {
            "host": "ebooking.ctrip.com",
            "pathName": "/promotion/promotionCenter",
            "locale": "zh-CN",
            "release": "",
            "client": {
                "deviceType": "PC",
                "os": "Windows",
                "osVersion": "Windows 10",
                "deviceName": "Windows PC",
                "clientId": "09031067317598779101",
                "screenWidth": 1536,
                "screenHeight": 864,
                "isIn": {
                    "ie": False,
                    "chrome": True,
                    "chrome49": False,
                    "wechat": False,
                    "firefox": False,
                    "ios": False,
                    "android": False,
                },
                "isModernBrowser": True,
                "browser": "Chrome",
                "browserVersion": "149",
                "platform": "pc",
                "technology": "web",
            },
            "ubt": {
                "pageid": "10650086872",
                "pvid": 27,
                "sid": 13,
                "vid": "1781678703564.c4ddQztFbxsZ",
                "fp": "A5E79A-950592-9CB598",
            },
            "gps": {"coord": "", "lat": "", "lng": "", "cid": 0, "cnm": ""},
            "protocal": "https:",
        },
        "promotionStatusList": ["EFFECTIVE", "AUDIT", "EXPIRED", "PAUSE"],
        "cipher": None,
        "head": {
            "cid": "09031067317598779101",
            "ctok": "",
            "cver": "1.0",
            "lang": "01",
            "sid": "8888",
            "syscode": "09",
            "auth": "",
            "xsid": "",
            "extension": [],
        },
    }


def parse_number(value: Any) -> Any:
    if value in (None, "", "-", "--"):
        return ""
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip().replace(",", "")
    try:
        number = float(text)
        return int(number) if number.is_integer() else number
    except ValueError:
        return value


def join_unique(values: list[Any], sep: str = "\uff1b") -> str:
    result: list[str] = []
    for value in values:
        text = str(value).strip() if value not in (None, "") else ""
        if text and text not in result:
            result.append(text)
    return sep.join(result)


def activity_source_type(activity: dict[str, Any]) -> str:
    return "\u5df2\u53c2\u4e0e\u4fc3\u9500"


def activity_status(activity: dict[str, Any]) -> str:
    raw = activity.get("promotionStatus")
    return STATUS_MAP.get(str(raw), str(raw or ""))


def discount_text(method: Any, percent: Any, value: Any) -> str:
    if percent not in (None, ""):
        if method == "DISCOUNT" and value not in (None, ""):
            try:
                return f"{float(value) * 10:g}\u6298"
            except (TypeError, ValueError):
                pass
        return f"\u7acb\u51cf {percent}%"
    return str(value or "")


def activity_time(activity: dict[str, Any]) -> str:
    ranges: list[str] = []
    for rule in activity.get("promotionRuleItemList") or []:
        start = rule.get("promotionStartDate")
        end = rule.get("promotionEndDate")
        if start or end:
            ranges.append(f"{start or ''} \u81f3 {end or ''}".strip())
    coupon_info = activity.get("couponAssembleInfo") or {}
    for item in coupon_info.get("couponInfoItems") or []:
        start = item.get("startDate")
        end = item.get("endDate")
        if start or end:
            ranges.append(f"{start or ''} \u81f3 {end or ''}".strip())
    return join_unique(ranges)


def activity_rule_labels(activity: dict[str, Any]) -> str:
    labels = [activity.get("promotionDesc")]
    campaign = activity.get("campaignAssembleInfo") or {}
    if campaign.get("promotionType"):
        labels.append(campaign.get("promotionType"))
    for rule in activity.get("promotionRuleItemList") or []:
        for detail in rule.get("promotionDetailItemList") or []:
            labels.append(discount_text(detail.get("promotionMethod") or rule.get("promotionMethod"), detail.get("pricePromotionValPercent"), detail.get("pricePromotionVal")))
            start_hour = detail.get("supportStartHour")
            end_hour = detail.get("supportEndHour")
            if start_hour not in (None, -1) or end_hour not in (None, -1):
                labels.append(f"{start_hour}:00-{end_hour}:00")
    coupon_info = activity.get("couponAssembleInfo") or {}
    for item in coupon_info.get("couponInfoItems") or []:
        labels.append(item.get("couponDiscountDesc"))
        if item.get("isAutoDelay"):
            labels.append("\u5230\u671f\u524d\u81ea\u52a8\u5ef6\u671f")
    return join_unique(labels)


def resource_rows(activity: dict[str, Any]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for rule in activity.get("promotionRuleItemList") or []:
        resource_item = rule.get("promotionResourceItem") or {}
        resource_type = resource_item.get("resourceDataType")
        resources = resource_item.get("resourceAssmbleItemList") or []
        if resource_type == "CHILD_HOTEL":
            rows.append({"id": resources[0].get("resourceId") if resources else "", "name": "\u5168\u623f\u578b", "inventory": "\u4e0d\u9650"})
            continue
        for item in resources:
            rows.append(
                {
                    "id": item.get("resourceId"),
                    "name": item.get("resourceName") or f"\u623f\u578b {item.get('resourceId')}",
                    "inventory": "\u4e0d\u9650",
                }
            )
    coupon_info = activity.get("couponAssembleInfo") or {}
    for item in coupon_info.get("couponInfoItems") or []:
        if item.get("resourceDataType") == "CHILD_HOTEL":
            rows.append({"id": (item.get("resourceIds") or [""])[0], "name": "\u5168\u623f\u578b", "inventory": "\u4e0d\u9650"})
        for resource in item.get("resourceInfoList") or []:
            rows.append({"id": resource.get("resourceId"), "name": "\u5168\u623f\u578b", "inventory": "\u4e0d\u9650"})
    unique: dict[tuple[str, str], dict[str, Any]] = {}
    for row in rows:
        unique[(str(row.get("id") or ""), str(row.get("name") or ""))] = row
    return list(unique.values())


def normalize_rows(payload: dict[str, Any], captured_at: datetime) -> tuple[list[list[Any]], list[list[Any]]]:
    summary_rows: list[list[Any]] = []
    detail_rows: list[list[Any]] = []
    for activity in payload.get("entities") or []:
        if not isinstance(activity, dict):
            continue
        source_type = activity_source_type(activity)
        name = activity.get("promotionName") or ""
        resources = resource_rows(activity)
        product_summary = join_unique([item.get("name") for item in resources])
        summary_rows.append(
            [
                captured_at,
                CHANNEL_SOURCE,
                source_type,
                name,
                activity_status(activity),
                activity_time(activity),
                activity_rule_labels(activity),
                product_summary,
            ]
        )
        for item in resources:
            detail_rows.append(
                [
                    captured_at,
                    CHANNEL_SOURCE,
                    source_type,
                    name,
                    item.get("id") or "",
                    item.get("name") or "",
                    parse_number(item.get("inventory")),
                ]
            )
    return summary_rows, detail_rows


def write_single_sheet(output_path: Path, sheet_name: str, headers: list[str], rows: list[list[Any]], widths: list[int]) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append(row)
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
    for row in ws.iter_rows(min_row=2):
        row[0].number_format = "yyyy-mm-dd hh:mm:ss"
    wb.save(output_path)
    return output_path


def json_safe(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return value


def write_standard_json(output_path: Path, headers: list[str], rows: list[list[Any]]) -> Path:
    payload = {
        "table_name": output_path.stem,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "row_count": len(rows),
        "rows": [{header: json_safe(row[index]) if index < len(row) else "" for index, header in enumerate(headers)} for row in rows],
    }
    json_path = output_path.with_suffix(".json")
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return json_path


def save_outputs(summary_rows: list[list[Any]], detail_rows: list[list[Any]], sync_db: bool = False) -> tuple[Path, Path]:
    summary_headers = [*SUMMARY_HEADERS, "hotel_id"]
    detail_headers = [*DETAIL_HEADERS, "hotel_id"]
    summary_rows = [list(row) + [HOTEL_ID] for row in summary_rows]
    detail_rows = [list(row) + [HOTEL_ID] for row in detail_rows]
    summary_path = write_single_sheet(PROMOTION_OUTPUT, SUMMARY_SHEET, summary_headers, summary_rows, [20, 10, 18, 28, 14, 24, 60, 60, 16])
    detail_path = write_single_sheet(DETAIL_OUTPUT, DETAIL_SHEET, detail_headers, detail_rows, [20, 10, 18, 28, 18, 60, 14, 16])
    write_standard_json(summary_path, summary_headers, summary_rows)
    write_standard_json(detail_path, detail_headers, detail_rows)
    if sync_db:
        sync_table(summary_path.stem, summary_headers, summary_rows, allow_empty_replace=True)
        sync_table(detail_path.stem, detail_headers, detail_rows, allow_empty_replace=True)
    return summary_path, detail_path


def main() -> None:
    parser = argparse.ArgumentParser(description="\u643a\u7a0b\u6d3b\u52a8\u53ca\u6298\u6263\u91c7\u96c6")
    parser.add_argument("--cookie", default=COOKIE)
    parser.add_argument("--promotion-url", default=GET_PRO_BATCH_URL)
    parser.add_argument("--sync-db", action="store_true", help="\u540c\u6b65\u5199\u5165 MySQL\uff1b\u9ed8\u8ba4\u53ea\u751f\u6210 Excel/JSON")
    args = parser.parse_args()
    captured_at = datetime.now()
    payload = CtripPromotionClient(args.cookie).query_promotion_batch(args.promotion_url)
    summary_rows, detail_rows = normalize_rows(payload, captured_at)
    summary_path, detail_path = save_outputs(summary_rows, detail_rows, sync_db=args.sync_db)
    print(f"OK \u643a\u7a0b\u6d3b\u52a8\u6570={len(summary_rows)} \u8f93\u51fa={summary_path}")
    print(f"OK \u643a\u7a0b\u6d3b\u52a8\u4ea7\u54c1\u660e\u7ec6\u6570={len(detail_rows)} \u8f93\u51fa={detail_path}")


if __name__ == "__main__":
    main()
