from __future__ import annotations

import argparse
import json
import os
import re
import time
from datetime import date, datetime
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from ctrip_config import COOKIE, DEFAULT_HOTEL_NAME, EXTRA_HEADERS, PLATFORM_SCOPE, USER_AGENT

import sys

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from ota_mysql_writer import DB_CONFIG, OUTPUT_DIR


OUTPUT_PATH = OUTPUT_DIR / "ctrip_ota_review_detail.xlsx"
TABLE_NAME = "ctrip_ota_review_detail"
API_URL = os.environ.get(
    "CTRIP_GET_COMMENT_LIST_URL",
    "https://ebooking.ctrip.com/restapi/soa2/26353/getCommentList"
    "?_fxpcqlniredt=09031067317598779101",
).strip()
HOTEL_NAME = os.environ.get("CTRIP_HOTEL_NAME", DEFAULT_HOTEL_NAME).strip()
HOTEL_ID = os.environ.get("CTRIP_HOTEL_ID", "").strip()
INTERNAL_HOTEL_ID = os.environ.get("HOTEL_ID", "").strip()

HEADERS = [
    "snapshot_time", "channel_source", "hotel_name", "poi_id", "review_id",
    "reviewer_name_masked", "review_score", "review_content", "review_time",
    "stay_date", "merchant_reply_content", "merchant_reply_time", "is_replied",
    "room_type_name", "ota_product_name", "has_image", "image_count",
    "image_urls_json", "is_anonymous", "is_negative_review", "read_status",
    "hygiene_score", "facility_score", "location_score", "service_score",
]


class CtripReviewDetailError(RuntimeError):
    pass


class CtripReviewDetailClient:
    def __init__(self, cookie: str):
        self.session = requests.Session()
        self.session.headers.update({
            "User-Agent": USER_AGENT,
            "Accept": "application/json, text/plain, */*",
            "Content-Type": "application/json;charset=UTF-8",
            "Origin": "https://ebooking.ctrip.com",
            "Referer": "https://ebooking.ctrip.com/comment/commentList?microJump=true",
        })
        if cookie.strip():
            self.session.headers["Cookie"] = cookie.strip()
        if EXTRA_HEADERS:
            self.session.headers.update(EXTRA_HEADERS)

    def fetch_page(self, page_index: int, page_size: int) -> dict[str, Any]:
        response = self.session.post(
            API_URL,
            json=build_payload(page_index, page_size),
            timeout=30,
        )
        response.raise_for_status()
        try:
            payload = response.json()
        except ValueError as exc:
            raise CtripReviewDetailError("Comment list response is not JSON") from exc
        if not isinstance(payload, dict):
            raise CtripReviewDetailError("Comment list response is not an object")
        ack = (payload.get("ResponseStatus") or {}).get("Ack")
        if ack not in (None, "Success"):
            raise CtripReviewDetailError(f"Comment list failed: Ack={ack}")
        status = payload.get("resStatus") or {}
        code = status.get("rcode") if isinstance(status, dict) else None
        if code not in (None, 0, "0", 200, "200"):
            raise CtripReviewDetailError(f"Comment list failed: rcode={code}")
        return payload


def build_payload(page_index: int, page_size: int) -> dict[str, Any]:
    return {
        "reqHead": {
            "host": "ebooking.ctrip.com",
            "pathName": "/comment/commentList",
            "locale": "zh-CN",
            "release": "",
            "client": {
                "deviceType": "PC", "os": "Windows", "osVersion": "Windows 10",
                "deviceName": "Windows PC", "clientId": "09031067317598779101",
                "screenWidth": 2560, "screenHeight": 1440,
                "isIn": {"ie": False, "chrome": True, "chrome49": False,
                         "wechat": False, "firefox": False, "ios": False, "android": False},
                "isModernBrowser": True, "browser": "Chrome", "browserVersion": "149",
                "platform": "pc", "technology": "web",
            },
            "ubt": {"pageid": "10650085973", "pvid": 4, "sid": 17,
                    "vid": "1781678703564.c4ddQztFbxsZ", "fp": ""},
            "gps": {"coord": "", "lat": "", "lng": "", "cid": 0, "cnm": ""},
            "protocal": "https:",
        },
        "keyWord": "", "pageIndex": page_index, "commentStatus": "",
        "isNeedTranslate": False, "sortType": 0, "catalogTab": "all",
        "catalogName": "\u5168\u90e8\u70b9\u8bc4", "pageSize": page_size,
        "needOrder": True, "endDate": "", "startDate": "",
        "channelSource": "1", "header": {"platform": "WEB"},
    }


def nested(item: dict[str, Any], *keys: str, default: Any = "") -> Any:
    for key in keys:
        value: Any = item
        for part in key.split("."):
            value = value.get(part) if isinstance(value, dict) else None
        if value not in (None, ""):
            return value
    return default


def to_datetime(value: Any) -> datetime | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    text = str(value).strip()
    match = re.search(r"/Date\((\d+)", text)
    if match:
        text = match.group(1)
    if text.replace(".", "", 1).isdigit():
        number = float(text)
        if number > 10_000_000_000:
            number /= 1000
        return datetime.fromtimestamp(number)
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S",
                "%Y-%m-%d", "%Y/%m/%d", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(text[:19], fmt)
        except ValueError:
            continue
    return None


def to_date(value: Any) -> date | None:
    parsed = to_datetime(value)
    return parsed.date() if parsed else None


def normalize_score(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if number > 10:
        number /= 10
    elif number > 5:
        number /= 2
    return round(number, 2)


def normalize_read_status(value: Any) -> int | None:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return int(value)
    text = str(value).strip().lower()
    if text in {"true", "read", "yes", "已读"}:
        return 1
    if text in {"false", "unread", "no", "未读"}:
        return 0
    try:
        return int(text)
    except ValueError:
        return None


def find_comment_items(payload: Any) -> list[dict[str, Any]]:
    candidates: list[tuple[int, list[dict[str, Any]]]] = []
    keys = {"reviewId", "commentId", "id", "content", "commentContent", "commentDate"}

    def walk(value: Any, name: str = "") -> None:
        if isinstance(value, dict):
            for child_name, child in value.items():
                walk(child, child_name)
        elif isinstance(value, list) and value and isinstance(value[0], dict):
            score = len(keys.intersection(value[0]))
            if name.lower() in {"commentlist", "comments", "items", "list"}:
                score += 5
            if score:
                candidates.append((score, [x for x in value if isinstance(x, dict)]))

    walk(payload)
    return max(candidates, key=lambda pair: pair[0])[1] if candidates else []


def image_urls(item: dict[str, Any]) -> list[str]:
    images = nested(item, "imageList", "images", "pictures", "picList", "picUrls", default=[])
    if isinstance(images, str):
        return [images] if images else []
    result: list[str] = []
    for image in images if isinstance(images, list) else []:
        if isinstance(image, str):
            result.append(image)
        elif isinstance(image, dict):
            url = nested(image, "url", "imageUrl", "originUrl", "bigUrl", "src")
            if url:
                result.append(str(url))
    return result


def dimension_scores(item: dict[str, Any]) -> dict[str, float | None]:
    result: dict[str, float | None] = {}
    values = nested(item, "scoreList", "subScores", "dimensionScores", "ratings", default=[])
    for score in values if isinstance(values, list) else []:
        if not isinstance(score, dict):
            continue
        name = str(nested(score, "title", "name", "type", "dimensionName"))
        value = normalize_score(nested(score, "score", "value", "scoreSimple"))
        for keyword, field in (("\u536b\u751f", "hygiene"), ("\u623f\u95f4", "hygiene"),
                               ("\u8bbe\u65bd", "facility"), ("\u73af\u5883", "location"),
                               ("\u4f4d\u7f6e", "location"), ("\u670d\u52a1", "service")):
            if keyword in name:
                result[field] = value
    return result


def normalize_rows(payload: dict[str, Any], captured_at: datetime | None = None) -> list[list[Any]]:
    captured_at = captured_at or datetime.now()
    rows: list[list[Any]] = []
    for item in find_comment_items(payload):
        review_id = str(nested(item, "reviewId", "commentId", "id", "commentID")).strip()
        if not review_id:
            continue
        images = image_urls(item)
        scores = dimension_scores(item)
        reply = str(nested(item, "replyContent", "hotelReply", "merchantReply",
                           "replyInfo.content", "reply.content")).strip()
        hotel_id = str(nested(item, "hotelId", "hotelID", default=HOTEL_ID)).strip()
        rating = normalize_score(nested(item, "reviewScore", "score", "rating", "totalScore"))
        negative = nested(item, "isNegative", "isBadComment", "badComment", default=None)
        rows.append([
            captured_at, "\u643a\u7a0b",
            str(nested(item, "hotelName", default=HOTEL_NAME)).strip(),
            hotel_id, review_id,
            str(nested(item, "userName", "userNickName", "nickName", "userInfo.nickName")).strip(),
            rating,
            str(nested(item, "content", "commentContent", "comment", "reviewContent")).strip(),
            to_datetime(nested(item, "commentTime", "commentDate", "createTime", "reviewTime")),
            to_date(nested(item, "checkInDate", "consumeTime", "stayDate", "orderInfo.checkInDate")),
            reply,
            to_datetime(nested(item, "replyTime", "replyDate", "replyInfo.replyTime", "reply.createTime")),
            bool(reply),
            str(nested(item, "roomName", "roomTypeName", "orderInfo.roomName")).strip(),
            str(nested(item, "productName", "ratePlanName", "orderName", "orderInfo.productName")).strip(),
            bool(images), len(images), json.dumps(images, ensure_ascii=False),
            bool(nested(item, "anonymous", "isAnonymous", default=False)),
            bool(negative) if negative is not None else bool(rating is not None and rating <= 2),
            normalize_read_status(nested(item, "readStatus", "isRead", default=None)),
            scores.get("hygiene"), scores.get("facility"), scores.get("location"), scores.get("service"),
        ])
    return rows


def find_total(payload: Any) -> int:
    if isinstance(payload, dict):
        for key in ("totalCount", "total", "recordCount"):
            if key in payload:
                try:
                    return int(payload[key])
                except (TypeError, ValueError):
                    pass
        for value in payload.values():
            total = find_total(value)
            if total:
                return total
    return 0


def collect_online(page_size: int, max_pages: int, full_history: bool) -> list[list[Any]]:
    client = CtripReviewDetailClient(COOKIE)
    captured_at = datetime.now()
    rows_by_id: dict[str, list[Any]] = {}
    previous_ids: set[str] = set()
    page = 1
    while full_history or page <= max_pages:
        payload = client.fetch_page(page, page_size)
        page_rows = normalize_rows(payload, captured_at)
        page_ids = {str(row[4]) for row in page_rows}
        if not page_rows or page_ids == previous_ids:
            break
        rows_by_id.update({str(row[4]): row for row in page_rows})
        if find_total(payload) and len(rows_by_id) >= find_total(payload):
            break
        previous_ids = page_ids
        page += 1
        time.sleep(0.15)
    return sorted(rows_by_id.values(), key=lambda row: row[8] or datetime.min, reverse=True)


def save_outputs(headers: list[str], rows: list[list[Any]]) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "\u8bc4\u4ef7\u660e\u7ec6"
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    widths = [20, 18, 35, 14, 18, 24, 14, 45, 20, 14, 32, 22, 14,
              28, 28, 14, 14, 40, 16, 20, 14, 16, 16, 16, 16]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width
    for row in sheet.iter_rows(min_row=2):
        for index in (1, 9, 12):
            row[index - 1].number_format = "yyyy-mm-dd hh:mm:ss"
        row[9].number_format = "yyyy-mm-dd"
    sheet.row_dimensions[1].height = 32
    workbook.save(OUTPUT_PATH)
    payload = {
        "table_name": OUTPUT_PATH.stem,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "row_count": len(rows),
        "rows": [{key: json_value(row[index]) for index, key in enumerate(headers)} for row in rows],
    }
    OUTPUT_PATH.with_suffix(".json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def json_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    return value


def upsert_mysql(headers: list[str], rows: list[list[Any]]) -> None:
    if not rows:
        print("DB sync skipped: no review detail rows")
        return
    import pymysql
    columns = ", ".join(f"`{name}`" for name in headers)
    placeholders = ", ".join(["%s"] * len(headers))
    updates = ", ".join(
        f"`{name}`=VALUES(`{name}`)" for name in headers
        if name not in {"channel_source", "poi_id", "review_id"})
    sql = (f"INSERT INTO `{TABLE_NAME}` ({columns}) VALUES ({placeholders}) "
           f"ON DUPLICATE KEY UPDATE {updates}")
    connection = pymysql.connect(**DB_CONFIG, autocommit=False)
    try:
        with connection.cursor() as cursor:
            cursor.executemany(sql, rows)
        connection.commit()
    except Exception:
        connection.rollback()
        raise
    finally:
        connection.close()
    print(f"DB upserted: {TABLE_NAME} rows={len(rows)}")


def sample_payload() -> dict[str, Any]:
    return {"result": {"totalCount": 1, "commentList": [{
        "commentId": "ctrip-demo-1", "userNickName": "\u533f***", "score": 4.5,
        "content": "\u623f\u95f4\u5e72\u51c0\uff0c\u670d\u52a1\u5f88\u597d",
        "commentDate": "2026-06-28 10:30:00", "checkInDate": "2026-06-27",
        "replyContent": "\u611f\u8c22\u60a8\u7684\u597d\u8bc4",
        "replyTime": "2026-06-28 12:00:00", "roomName": "\u7b80\u81f4\u5927\u5e8a\u623f",
        "productName": "\u4e0d\u542b\u65e9", "imageList": [{"url": "https://example.test/1.jpg"}],
        "isAnonymous": True, "readStatus": 1,
        "scoreList": [{"title": "\u670d\u52a1", "score": 5}, {"title": "\u8bbe\u65bd", "score": 4}],
    }]}}


def main() -> None:
    parser = argparse.ArgumentParser(description="Collect Ctrip review details.")
    parser.add_argument("--input-json")
    parser.add_argument("--self-test", action="store_true")
    parser.add_argument("--page-size", type=int, default=10)
    parser.add_argument("--max-pages", type=int, default=10)
    parser.add_argument("--full-history", action="store_true")
    parser.add_argument("--sync-db", action="store_true")
    args = parser.parse_args()
    if args.self_test:
        rows = normalize_rows(sample_payload())
    elif args.input_json:
        payload = json.loads(Path(args.input_json).read_text(encoding="utf-8-sig"))
        rows = normalize_rows(payload)
    else:
        if not COOKIE:
            raise CtripReviewDetailError("CTRIP_COOKIE is empty")
        rows = collect_online(max(1, args.page_size), max(1, args.max_pages), args.full_history)
    headers = [*HEADERS, "platform_scope", "hotel_id"]
    rows = [list(row) + [PLATFORM_SCOPE, INTERNAL_HOTEL_ID] for row in rows]
    save_outputs(headers, rows)
    if args.sync_db:
        upsert_mysql(headers, rows)
    print(f"review_detail rows={len(rows)}")
    print(f"Excel saved: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
