from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from ctrip_config import COOKIE, EXTRA_HEADERS, USER_AGENT
from ctrip_review_dom_counts import (
    collect_review_counts,
    load_existing_counts,
    merge_counts,
)

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from ota_mysql_writer import DB_CONFIG, OUTPUT_DIR, sync_table


CHANNEL_SOURCE = "\u643a\u7a0b"
GET_HOTEL_RATING_URL = os.environ.get(
    "CTRIP_GET_HOTEL_RATING_URL",
    "https://ebooking.ctrip.com/restapi/soa2/26353/getHotelRating"
    "?_fxpcqlniredt=09031067317598779101"
    "&x-traceID=09031067317598779101-1782378921227-7521729",
).strip()
REVIEW_CAPTURE = OUTPUT_DIR / "ctrip_review_dom_capture.json"
OVERVIEW_OUTPUT = OUTPUT_DIR / "ctrip_ota_review_overview.xlsx"
RANKING_OUTPUT = OUTPUT_DIR / "ctrip_ota_review_ranking.xlsx"
HOTEL_ID = os.environ.get("HOTEL_ID", "").strip()

OVERVIEW_SHEET = "\u8bc4\u4ef7\u6982\u89c8"
RANKING_SHEET = "\u8bc4\u4ef7\u6392\u884c"

OVERVIEW_HEADERS = [
    "snapshot_time",
    "channel_source",
    "review_score",
    "review_score_max",
    "environment_score",
    "facility_score",
    "service_score",
    "hygiene_score",
    "total_review_count",
    "unreplied_review_count",
    "negative_review_count",
]

RANKING_HEADERS = [
    "snapshot_time",
    "channel_source",
    "ranking_type",
    "ranking_position",
    "rank_item_name",
    "rank_item_value",
]


class CtripApiError(RuntimeError):
    pass


class CtripClient:
    def __init__(self, cookie: str, base_url: str = "https://ebooking.ctrip.com"):
        self.base_url = base_url.rstrip("/")
        self.session = requests.Session()
        self.session.headers.update(
            {
                "User-Agent": USER_AGENT,
                "Accept": "application/json, text/plain, */*",
                "Content-Type": "application/json;charset=UTF-8",
                "Origin": "https://ebooking.ctrip.com",
                "Referer": "https://ebooking.ctrip.com/",
            }
        )
        if cookie.strip():
            self.session.headers["Cookie"] = cookie.strip()
        if EXTRA_HEADERS:
            self.session.headers.update(EXTRA_HEADERS)

    def post_json(self, path_or_url: str, payload: dict[str, Any] | None = None) -> dict[str, Any]:
        response = self.session.post(self._url(path_or_url), json=payload or {}, timeout=30)
        response.raise_for_status()
        try:
            data = response.json()
        except ValueError as exc:
            raise CtripApiError(f"接口没有返回 JSON，HTTP={response.status_code}") from exc
        if not isinstance(data, dict):
            raise CtripApiError("接口 JSON 顶层不是 object")
        return data

    def _url(self, path_or_url: str) -> str:
        if path_or_url.startswith("http://") or path_or_url.startswith("https://"):
            return path_or_url
        return f"{self.base_url}/{path_or_url.lstrip('/')}"


def require_config(name: str, value: str) -> str:
    if not value or "TODO" in value:
        raise CtripApiError(f"请先配置 {name}")
    return value


def unwrap_payload(data: dict[str, Any]) -> Any:
    code = data.get("code", data.get("status", data.get("resultCode")))
    success = data.get("success")
    if success is False or code not in (None, 0, "0", 200, "200", "success"):
        message = data.get("message") or data.get("msg") or data.get("errorMsg")
        raise CtripApiError(f"接口返回异常：code={code}, message={message}")
    for key in ("data", "result", "value"):
        if key in data:
            return data[key]
    return data


def parse_percent(value: Any) -> Any:
    if value is None or value == "":
        return ""
    if isinstance(value, (int, float)):
        return value if abs(value) <= 1 else value / 100
    text = str(value).strip().replace(",", "")
    if text in ("-", "--"):
        return "-"
    if text.endswith("%"):
        text = text[:-1]
    try:
        return float(text) / 100
    except ValueError:
        return value


def parse_number(value: Any) -> Any:
    if value is None or value == "":
        return ""
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip().replace(",", "")
    if text in ("-", "--"):
        return "-"
    if text.endswith("%"):
        return parse_percent(text)
    try:
        number = float(text)
        return int(number) if number.is_integer() else number
    except ValueError:
        return value


def write_single_sheet(
    output_path: Path,
    sheet_name: str,
    headers: list[str],
    rows: list[list[Any]],
    widths: list[int],
    datetime_columns: set[int] | None = None,
    percent_columns: set[int] | None = None,
) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append(row)
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if datetime_columns and cell.column in datetime_columns:
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
            if percent_columns and cell.column in percent_columns:
                cell.number_format = "0.00%"
    wb.save(output_path)
    return output_path


def json_safe(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if hasattr(value, "isoformat") and not isinstance(value, (str, int, float, bool)):
        return value.isoformat()
    return value


def write_standard_json(output_path: Path, headers: list[str], rows: list[list[Any]]) -> Path:
    json_path = output_path.with_suffix(".json")
    payload = {
        "table_name": output_path.stem,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "row_count": len(rows),
        "rows": [
            {header: json_safe(row[index]) if index < len(row) else "" for index, header in enumerate(headers)}
            for row in rows
        ],
    }
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return json_path


class CtripReviewClient(CtripClient):
    def query_hotel_rating(self, url: str = GET_HOTEL_RATING_URL) -> dict[str, Any]:
        data = self.post_json(require_config("GET_HOTEL_RATING_URL", url), build_rating_payload())
        response_status = data.get("ResponseStatus") or {}
        if isinstance(response_status, dict) and response_status.get("Ack") not in (None, "Success"):
            raise CtripApiError(f"getHotelRating 返回异常：{response_status.get('Ack')}")
        res_status = data.get("resStatus") or {}
        code = res_status.get("rcode") if isinstance(res_status, dict) else None
        if code not in (None, 0, "0", 200, "200"):
            raise CtripApiError(f"getHotelRating 返回异常：rcode={code}")
        return data


def build_rating_payload() -> dict[str, Any]:
    return {
        "reqHead": {
            "host": "ebooking.ctrip.com",
            "pathName": "/comment/commentList",
            "locale": "zh-CN",
            "release": "",
            "client": {
                "deviceType": "PC",
                "os": "Windows",
                "osVersion": "Windows 10",
                "deviceName": "Windows PC",
                "clientId": "09031067317598779101",
                "screenWidth": 1536,
                "screenHeight": 864,
                "isIn": {
                    "ie": False,
                    "chrome": True,
                    "chrome49": False,
                    "wechat": False,
                    "firefox": False,
                    "ios": False,
                    "android": False,
                },
                "isModernBrowser": True,
                "browser": "Chrome",
                "browserVersion": "149",
                "platform": "pc",
                "technology": "web",
            },
            "ubt": {
                "pageid": "10650085973",
                "pvid": 2,
                "sid": 13,
                "vid": "1781678703564.c4ddQztFbxsZ",
                "fp": "65879A-950592-EBA3F9",
            },
            "gps": {"coord": "", "lat": "", "lng": "", "cid": 0, "cnm": ""},
            "protocal": "https:",
        },
        "channelSource": "trip",
        "header": {"platform": "WEB"},
        "head": {
            "cid": "09031067317598779101",
            "ctok": "",
            "cver": "1.0",
            "lang": "01",
            "sid": "8888",
            "syscode": "09",
            "auth": "",
            "xsid": "",
            "extension": [],
        },
    }


def pick(item: dict[str, Any], *keys: str, default: Any = "") -> Any:
    for key in keys:
        if key in item and item[key] not in (None, ""):
            return item[key]
    return default


def find_card(cards: list[dict[str, Any]], *names: str) -> dict[str, Any]:
    for card in cards:
        title = str(pick(card, "title", "name", "metricName"))
        if any(name in title for name in names):
            return card
    return {}


def ext_value(card: dict[str, Any], *names: str) -> Any:
    attrs = card.get("extAttrs") or card.get("attrs") or card.get("subItems") or []
    if isinstance(attrs, dict):
        attrs = attrs.values()
    for attr in attrs:
        if not isinstance(attr, dict):
            continue
        name = str(pick(attr, "name", "title", "label"))
        if any(target in name for target in names):
            values = attr.get("values")
            if isinstance(values, list) and values:
                return values[0]
            return pick(attr, "value", "num")
    return ""


def card_value(card: dict[str, Any]) -> Any:
    return pick(card, "value", "metricValue", "num", "count")


def extract_cards(payload: Any) -> list[dict[str, Any]]:
    if isinstance(payload, list):
        return [item for item in payload if isinstance(item, dict)]
    if isinstance(payload, dict):
        for key in ("cards", "metrics", "items", "overview"):
            value = payload.get(key)
            if isinstance(value, list):
                return [item for item in value if isinstance(item, dict)]
    return []


def normalize_overview_rows(overviews: list[tuple[int, Any]], captured_at: datetime) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for period_days, payload in overviews:
        cards = extract_cards(payload)
        bad_new = find_card(cards, "新增差评", "差评数", "bad")
        bad_pending = find_card(cards, "待回复", "未回复")
        reply_rate = find_card(cards, "回复率")
        positive_new = find_card(cards, "新增好评", "好评数", "positive")
        rows.append(
            [
                captured_at,
                CHANNEL_SOURCE,
                period_days,
                parse_number(card_value(bad_new)),
                parse_number(ext_value(bad_new, "有内容", "content")),
                parse_number(ext_value(bad_new, "带图片", "image", "图片")),
                parse_number(card_value(bad_pending)),
                parse_number(ext_value(bad_pending, "有内容", "content")),
                parse_number(ext_value(bad_pending, "带图片", "image", "图片")),
                parse_percent(card_value(reply_rate)),
                parse_percent(ext_value(reply_rate, "同行", "标杆", "peer")),
                parse_number(card_value(positive_new)),
                parse_number(ext_value(positive_new, "有内容", "content")),
                parse_number(ext_value(positive_new, "带图片", "image", "图片")),
            ]
        )
    return rows


def extract_rank_items(payload: Any, key: str) -> list[dict[str, Any]]:
    if isinstance(payload, dict):
        value = payload.get(key)
        if isinstance(value, list):
            return [item for item in value if isinstance(item, dict)]
    return []


def normalize_ranking_rows(payload: Any, captured_at: datetime) -> list[list[Any]]:
    mapping = {"positive": "好评热词", "negative": "差评热词", "peer": "同行评分"}
    rows: list[list[Any]] = []
    for key, label in mapping.items():
        for index, item in enumerate(extract_rank_items(payload, key), 1):
            keyword = pick(item, "keyword", "name", "title")
            if isinstance(keyword, list):
                keyword = "、".join(str(x) for x in keyword)
            rows.append(
                [
                    captured_at,
                    CHANNEL_SOURCE,
                    label,
                    pick(item, "position", "rank", default=index),
                    keyword,
                    parse_number(pick(item, "number", "value", "score")),
                ]
            )
    return rows


def load_review_capture(path: Path = REVIEW_CAPTURE) -> dict[str, Any]:
    if not path.exists():
        raise CtripApiError(f"未找到页面快照文件：{path}")
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise CtripApiError(f"页面快照读取失败：{path}") from exc
    if not isinstance(data, dict):
        raise CtripApiError("页面快照顶层不是 object")
    return data


def normalize_capture_overview_rows(capture: dict[str, Any], captured_at: datetime) -> list[list[Any]]:
    overview = capture.get("overview") or {}
    if not isinstance(overview, dict):
        overview = {}
    channel_source = overview.get("channelSource") or CHANNEL_SOURCE
    return [
        [
            captured_at,
            channel_source,
            parse_number(overview.get("reviewScore")),
            parse_number(overview.get("reviewScoreMax")),
            parse_number(overview.get("environmentScore")),
            parse_number(overview.get("facilityScore")),
            parse_number(overview.get("serviceScore")),
            parse_number(overview.get("hygieneScore")),
            parse_number(overview.get("totalReviewCount")),
            parse_number(overview.get("pendingReplyCount")),
            parse_number(overview.get("badReviewCount")),
        ]
    ]


def normalize_capture_ranking_rows(capture: dict[str, Any], captured_at: datetime) -> list[list[Any]]:
    overview = capture.get("overview") or {}
    impressions = capture.get("impressions") or {}
    if not isinstance(overview, dict):
        overview = {}
    if not isinstance(impressions, dict):
        impressions = {}
    channel_source = overview.get("channelSource") or CHANNEL_SOURCE
    rows: list[list[Any]] = []
    for key, label in (("positive", "positive_impression"), ("negative", "negative_impression")):
        items = impressions.get(key) or []
        if not isinstance(items, list):
            continue
        for index, item in enumerate(items, 1):
            if not isinstance(item, dict):
                continue
            rows.append(
                [
                    captured_at,
                    channel_source,
                    label,
                    index,
                    item.get("keyword", ""),
                    parse_number(item.get("value")),
                ]
            )
    return rows


def rating_info(payload: dict[str, Any]) -> dict[str, Any]:
    info = payload.get("ratingInfo") or payload.get("ctripRatings") or {}
    return info if isinstance(info, dict) else {}


def score_from_sub_scores(info: dict[str, Any], scalar_key: str, sub_type: str) -> Any:
    if info.get(scalar_key) not in (None, ""):
        return parse_number(info.get(scalar_key))
    score_info = info.get("scoreInfo") or {}
    if not isinstance(score_info, dict):
        return ""
    sub_scores = score_info.get("subScores") or []
    if not isinstance(sub_scores, list):
        return ""
    for item in sub_scores:
        if not isinstance(item, dict) or item.get("type") != sub_type:
            continue
        return parse_number(item.get("scoreSimple") or item.get("score"))
    return ""


def normalize_rating_overview_rows(payload: dict[str, Any], captured_at: datetime) -> list[list[Any]]:
    info = rating_info(payload)
    score_info = info.get("scoreInfo") or {}
    if not isinstance(score_info, dict):
        score_info = {}
    return [
        [
            captured_at,
            CHANNEL_SOURCE,
            parse_number(info.get("ratingAll") or score_info.get("avgScoreSimple") or score_info.get("avgScore")),
            parse_number(score_info.get("maxScore") or 5),
            score_from_sub_scores(info, "ratingLocation", "ratingLocation"),
            score_from_sub_scores(info, "ratingFacility", "ratingFacility"),
            score_from_sub_scores(info, "ratingService", "ratingService"),
            score_from_sub_scores(info, "ratingRoom", "ratingRoom"),
            "",
            "",
            "",
        ]
    ]


def normalize_rating_ranking_rows(payload: dict[str, Any], captured_at: datetime) -> list[list[Any]]:
    info = rating_info(payload)
    rows: list[list[Any]] = []
    for key, label in (("goodCommentTags", "positive_impression"), ("poorCommentTags", "negative_impression")):
        items = info.get(key) or []
        if not isinstance(items, list):
            continue
        for index, item in enumerate(items, 1):
            if not isinstance(item, dict):
                continue
            rows.append(
                [
                    captured_at,
                    CHANNEL_SOURCE,
                    label,
                    index,
                    item.get("tagName") or item.get("name") or "",
                    parse_number(item.get("tagCount") or item.get("count")),
                ]
            )
    return rows


def sample_capture() -> dict[str, Any]:
    return {
        "overview": {
            "channelSource": CHANNEL_SOURCE,
            "reviewScore": "4.4",
            "reviewScoreMax": "5",
            "environmentScore": "4.4",
            "facilityScore": "4.3",
            "serviceScore": "4.5",
            "hygieneScore": "4.4",
            "totalReviewCount": "922",
            "pendingReplyCount": "0",
            "badReviewCount": "22",
        },
        "impressions": {
            "positive": [{"keyword": "环境优雅", "value": 84}],
            "negative": [{"keyword": "设施一般", "value": 7}],
        },
    }


def sample_overview() -> list[tuple[int, Any]]:
    card = lambda title, value: {"title": title, "value": value, "extAttrs": [{"name": "有内容", "values": [value]}, {"name": "带图片", "values": ["0"]}]}
    return [
        (1, {"cards": [card("新增差评数", "0"), card("待回复差评数", "0"), {"title": "差评回复率", "value": "100.00%", "extAttrs": [{"name": "同行标杆", "values": ["95.00%"]}]}, card("新增好评数", "8")]}),
        (7, {"cards": [card("新增差评数", "1"), card("待回复差评数", "0"), {"title": "差评回复率", "value": "100.00%", "extAttrs": [{"name": "同行标杆", "values": ["95.00%"]}]}, card("新增好评数", "26")]}),
        (30, {"cards": [card("新增差评数", "2"), card("待回复差评数", "1"), {"title": "差评回复率", "value": "50.00%", "extAttrs": [{"name": "同行标杆", "values": ["95.00%"]}]}, card("新增好评数", "112")]}),
    ]


def sample_ranking() -> dict[str, Any]:
    return {
        "positive": [{"position": 1, "keyword": ["房间好"], "number": "54"}],
        "negative": [{"position": 1, "keyword": ["设施一般"], "number": "2"}],
        "peer": [{"position": 1, "keyword": ["示例酒店"], "number": "4.9"}],
    }


def save_outputs(overview_rows: list[list[Any]], ranking_rows: list[list[Any]], sync_db: bool = False) -> tuple[Path, Path]:
    overview_headers = [*OVERVIEW_HEADERS, "hotel_id"]
    ranking_headers = [*RANKING_HEADERS, "hotel_id"]
    overview_rows = [list(row) + [HOTEL_ID] for row in overview_rows]
    ranking_rows = [list(row) + [HOTEL_ID] for row in ranking_rows]
    overview = write_single_sheet(
        OVERVIEW_OUTPUT,
        OVERVIEW_SHEET,
        overview_headers,
        overview_rows,
        widths=[20, 12, 14, 14, 16, 16, 16, 16, 18, 18, 16, 16],
        datetime_columns={1},
    )
    ranking = write_single_sheet(
        RANKING_OUTPUT,
        RANKING_SHEET,
        ranking_headers,
        ranking_rows,
        widths=[20, 10, 14, 10, 34, 14, 16],
        datetime_columns={1},
    )
    write_standard_json(overview, overview_headers, overview_rows)
    write_standard_json(ranking, ranking_headers, ranking_rows)
    if sync_db:
        sync_table(overview.stem, overview_headers, overview_rows)
        sync_table(ranking.stem, ranking_headers, ranking_rows)
    return overview, ranking


def main() -> None:
    parser = argparse.ArgumentParser(description="携程评价数据采集框架")
    parser.add_argument("--self-test", action="store_true")
    parser.add_argument("--cookie", default=COOKIE)
    parser.add_argument("--rating-url", default=GET_HOTEL_RATING_URL)
    parser.add_argument("--sync-db", action="store_true", help="同步写入 MySQL；默认只生成 Excel/JSON")
    args = parser.parse_args()
    captured_at = datetime.now()
    if args.self_test:
        capture = sample_capture()
        overview_rows = normalize_capture_overview_rows(capture, captured_at)
        ranking_rows = normalize_capture_ranking_rows(capture, captured_at)
    else:
        payload = CtripReviewClient(args.cookie).query_hotel_rating(args.rating_url)
        overview_rows = normalize_rating_overview_rows(payload, captured_at)
        ranking_rows = normalize_rating_ranking_rows(payload, captured_at)
        try:
            counts = collect_review_counts(
                args.cookie,
                expected_hotel_name=os.environ.get("CTRIP_HOTEL_NAME", ""),
                expected_hotel_id=os.environ.get("CTRIP_HOTEL_ID", ""),
            )
            print(
                "DOM review counts: "
                f"total={counts['total_review_count']} "
                f"unreplied={counts['unreplied_review_count']} "
                f"negative={counts['negative_review_count']}"
            )
        except Exception as exc:
            print(f"DOM review counts warning: {exc}")
            counts = load_existing_counts(DB_CONFIG, HOTEL_ID)
        merge_counts(overview_rows, counts)
    overview, ranking = save_outputs(overview_rows, ranking_rows, sync_db=args.sync_db)
    print(f"OK 携程评价概览行数={len(overview_rows)} 输出={overview}")
    print(f"OK 携程评价排行行数={len(ranking_rows)} 输出={ranking}")


if __name__ == "__main__":
    main()
