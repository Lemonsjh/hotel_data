from __future__ import annotations

import json
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def save_outputs(output_path: Path, headers: list[str], rows: list[list[Any]]) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "\u8bc4\u4ef7\u660e\u7ec6"
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    widths = [20, 18, 35, 14, 18, 24, 14, 45, 20, 14, 32, 22, 14, 28, 28, 14, 14, 40, 16, 20, 14, 16, 16, 16, 16]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width
    for row in sheet.iter_rows(min_row=2):
        for index in (1, 9, 12):
            row[index - 1].number_format = "yyyy-mm-dd hh:mm:ss"
        row[9].number_format = "yyyy-mm-dd"
    sheet.row_dimensions[1].height = 32
    workbook.save(output_path)
    payload = {
        "table_name": output_path.stem,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "row_count": len(rows),
        "rows": [{key: json_value(row[index]) for index, key in enumerate(headers)} for row in rows],
    }
    output_path.with_suffix(".json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def json_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    return value


def upsert_mysql(table_name: str, db_config: dict[str, Any], headers: list[str], rows: list[list[Any]]) -> None:
    if not rows:
        print("DB sync skipped: no review detail rows")
        return
    import pymysql

    columns = ", ".join(f"`{name}`" for name in headers)
    placeholders = ", ".join(["%s"] * len(headers))
    updates = ", ".join(f"`{name}`=VALUES(`{name}`)" for name in headers if name not in {"channel_source", "poi_id", "review_id"})
    sql = f"INSERT INTO `{table_name}` ({columns}) VALUES ({placeholders}) ON DUPLICATE KEY UPDATE {updates}"
    connection = pymysql.connect(**db_config, autocommit=False)
    try:
        with connection.cursor() as cursor:
            cursor.executemany(sql, rows)
            cleanup_legacy_blank_poi(cursor, table_name, headers, rows)
        connection.commit()
    except Exception:
        connection.rollback()
        raise
    finally:
        connection.close()
    print(f"DB upserted: {table_name} rows={len(rows)}")


def cleanup_legacy_blank_poi(cursor: Any, table_name: str, headers: list[str], rows: list[list[Any]]) -> None:
    required = {"channel_source", "platform_scope", "poi_id", "review_id"}
    if not required.issubset(headers):
        return
    positions = {name: headers.index(name) for name in required}
    groups: dict[tuple[str, str], set[str]] = {}
    for row in rows:
        poi_id = str(row[positions["poi_id"]] or "").strip()
        review_id = str(row[positions["review_id"]] or "").strip()
        if poi_id and review_id:
            key = (str(row[positions["channel_source"]]), str(row[positions["platform_scope"]]))
            groups.setdefault(key, set()).add(review_id)
    for (channel_source, platform_scope), review_ids in groups.items():
        marks = ", ".join(["%s"] * len(review_ids))
        cursor.execute(
            f"DELETE FROM `{table_name}` WHERE channel_source=%s AND platform_scope=%s AND poi_id='' AND review_id IN ({marks})",
            (channel_source, platform_scope, *sorted(review_ids)),
        )
