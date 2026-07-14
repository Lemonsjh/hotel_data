from __future__ import annotations

import argparse
import json
import os
import re
import sys
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import urlencode

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from meituan_config import MEITUAN_EB_COOKIE, MEITUAN_ME_COOKIE, PARTNER_ID, POI_ID, USER_AGENT

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from ota_mysql_writer import OUTPUT_DIR, sync_table


DEFAULT_EXCEL_PATH = OUTPUT_DIR / "meituan_ota_collected_data.xlsx"
SPLIT_OUTPUT_DIR = OUTPUT_DIR
OVERVIEW_SHEET = "\u8bc4\u4ef7\u6982\u89c8"
RANKING_SHEET = "\u8bc4\u4ef7\u6392\u884c"
PLATFORM = "\u7f8e\u56e2"


def build_ranking_url() -> str:
    query = {
        "poiId": POI_ID,
        "partnerId": PARTNER_ID,
        "source": 1,
        "yodaReady": "h5",
        "csecplatform": 4,
        "csecversion": "4.2.4",
    }
    return f"https://eb.meituan.com/api/shepherdGw/bizDatacenter/hotel/eb/dataCenter/service/ranking?{urlencode(query)}"


CONTRAST_URL = os.environ.get("MEITUAN_REVIEW_CONTRAST_URL", "").strip()
DIANPING_CONTRAST_URL = os.environ.get("MEITUAN_DIANPING_REVIEW_CONTRAST_URL", "").strip()
RANKING_URL = os.environ.get("MEITUAN_REVIEW_RANKING_URL", build_ranking_url()).strip()
HOTEL_ID = os.environ.get("HOTEL_ID", "").strip()

OVERVIEW_HEADERS = [
    "snapshot_time",
    "channel_source",
    "review_platform",
    "review_score",
    "review_score_max",
    "environment_score",
    "facility_score",
    "service_score",
    "hygiene_score",
    "total_review_count",
    "unreplied_review_count",
    "negative_review_count",
]

RANKING_HEADERS = [
    "snapshot_time",
    "channel_source",
    "ranking_type",
    "ranking_position",
    "rank_item_name",
    "rank_item_value",
]


class MeituanReviewError(RuntimeError):
    pass


class MeituanReviewClient:
    def __init__(self, cookie: str, contrast_url: str = CONTRAST_URL):
        self.contrast_url = contrast_url
        self.session = requests.Session()
        self.session.headers.update(
            {
                "User-Agent": USER_AGENT,
                "Accept": "application/json, text/plain, */*",
                "Referer": "https://me.meituan.com/ebooking/merchant/comment-manage-react",
                "Origin": "https://me.meituan.com",
                "X-Requested-With": "XMLHttpRequest",
                "Request-Page-Source": "ME",
            }
        )
        if cookie.strip():
            self.session.headers["Cookie"] = cookie.strip()

    def get_contrast(self, contrast_url: str, source_name: str) -> dict[str, Any]:
        if not contrast_url:
            raise MeituanReviewError(
                f"{source_name} contrast URL is empty; configure a current signed URL"
            )
        response = self.session.get(contrast_url, timeout=30)
        response.raise_for_status()
        response.encoding = "utf-8"
        data = response.json()
        if data.get("code") not in (0, 10000, "0", "10000"):
            msg = data.get("message") or (data.get("error") or {}).get("msg")
            raise MeituanReviewError(f"{source_name} contrast failed: code={data.get('code')}, msg={msg}")
        payload = data.get("data")
        if not isinstance(payload, dict):
            raise MeituanReviewError("contrast data is not object")
        return payload

    def get_contrast_new(self) -> dict[str, Any]:
        return self.get_contrast(self.contrast_url, "Meituan")

    def get_dianping_contrast(self) -> dict[str, Any]:
        return self.get_contrast(DIANPING_CONTRAST_URL, "Dianping")

    def get_ranking(self, ranking_url: str = RANKING_URL, cookie: str = MEITUAN_EB_COOKIE) -> dict[str, Any]:
        session = requests.Session()
        session.headers.update(
            {
                "User-Agent": USER_AGENT,
                "Accept": "application/json, text/plain, */*",
                "Referer": "https://eb.meituan.com/ebooking/hotel/dataCenter",
                "Origin": "https://eb.meituan.com",
                "X-Requested-With": "XMLHttpRequest",
            }
        )
        if cookie.strip():
            session.headers["Cookie"] = cookie.strip()
        response = session.get(ranking_url, timeout=30)
        response.raise_for_status()
        response.encoding = "utf-8"
        data = response.json()
        if data.get("status") not in (0, "0") and data.get("code") not in (0, 10000, "0", "10000"):
            raise MeituanReviewError(f"ranking failed: status={data.get('status')}, message={data.get('message')}")
        payload = data.get("data")
        if not isinstance(payload, dict):
            raise MeituanReviewError("ranking data is not object")
        return payload


def to_number(value: Any) -> Any:
    if value is None or value == "":
        return ""
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip().replace(",", "")
    if text in ("-", "--"):
        return "-"
    if text.endswith("%"):
        return to_percent(text)
    try:
        number = float(text)
        return int(number) if number.is_integer() else number
    except ValueError:
        return value


def to_percent(value: Any) -> Any:
    if value is None or value == "":
        return ""
    if isinstance(value, (int, float)):
        return value if abs(value) <= 1 else value / 100
    text = str(value).strip().replace(",", "")
    if text.endswith("%"):
        text = text[:-1]
    try:
        return float(text) / 100
    except ValueError:
        return value


def html_to_text(value: Any) -> str:
    text = "" if value is None else str(value)
    text = re.sub(r"<[^>]+>", "", text)
    return text.strip()


def sub_score(payload: dict[str, Any], title: str) -> Any:
    for item in payload.get("subScores") or []:
        if isinstance(item, dict) and item.get("title") == title:
            return to_number(item.get("score"))
    return ""


def first_present(*values: Any) -> Any:
    for value in values:
        if value is not None and value != "":
            return value
    return ""


def bad_count(total: Any, bad_rate: Any) -> Any:
    if total in (None, "") or bad_rate in (None, ""):
        return ""
    try:
        return round(float(total) * float(bad_rate))
    except (TypeError, ValueError):
        return ""


def normalize_overview_rows(
    payload: dict[str, Any], captured_at: datetime, review_platform: str = "meituan"
) -> list[list[Any]]:
    info = payload.get("commentAnalysisInfo") or {}
    if review_platform == "dianping":
        total = to_number(info.get("dpTotalCount30"))
        negative = to_number(info.get("dpBadCount30"))
    else:
        total = to_number(first_present(info.get("totalCount365"), info.get("totalCount180"), info.get("totalCount30")))
        bad_rate = to_percent(first_present(info.get("badPercent365"), info.get("badPercent180"), info.get("badPercent30")))
        negative = bad_count(total, bad_rate)
    return [
        [
            captured_at,
            PLATFORM,
            review_platform,
            (to_number(payload.get("score")) / 10) if payload.get("score") not in (None, "") else "",
            5,
            sub_score(payload, "\u4f4d\u7f6e"),
            sub_score(payload, "\u8bbe\u65bd"),
            sub_score(payload, "\u670d\u52a1"),
            sub_score(payload, "\u536b\u751f"),
            total,
            "",
            negative,
        ]
    ]


def normalize_ranking_rows(payload: dict[str, Any], captured_at: datetime) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for index, item in enumerate(payload.get("tagPos") or [], 1):
        if not isinstance(item, dict):
            continue
        rows.append(
            [
                captured_at,
                PLATFORM,
                "positive_impression" if item.get("type") == 1 else "guest_impression",
                index,
                item.get("tagName") or item.get("searchTag") or "",
                to_number(item.get("count")),
            ]
        )
    return rows


def normalize_keyword_ranking_rows(payload: dict[str, Any], captured_at: datetime) -> list[list[Any]]:
    rows: list[list[Any]] = []
    mapping = {
        "positive": "positive_keyword",
        "negative": "negative_keyword",
        "peer": "peer_score",
    }
    for key, ranking_type in mapping.items():
        for item in payload.get(key) or []:
            if not isinstance(item, dict):
                continue
            keyword = item.get("keyword")
            if isinstance(keyword, list):
                name = " ".join(str(value) for value in keyword if value not in (None, ""))
            else:
                name = to_text(keyword)
            rows.append(
                [
                    captured_at,
                    PLATFORM,
                    ranking_type,
                    to_number(item.get("position")),
                    name,
                    to_number(item.get("number")),
                ]
            )
    return rows


def get_or_create_workbook(path: Path) -> Workbook:
    if path.exists():
        return load_workbook(path)
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    return wb


def reset_sheet(wb: Workbook, sheet_name: str, headers: list[str], widths: list[int]):
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    ws.append(headers)
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
    return ws


def json_safe(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if hasattr(value, "isoformat") and not isinstance(value, (str, int, float, bool)):
        return value.isoformat()
    return value


def write_standard_json(output_path: Path, headers: list[str], rows: list[list[Any]]) -> Path:
    json_path = output_path.with_suffix(".json")
    payload = {
        "table_name": output_path.stem,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "row_count": len(rows),
        "rows": [
            {header: json_safe(row[index]) if index < len(row) else "" for index, header in enumerate(headers)}
            for row in rows
        ],
    }
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return json_path


def save_single_sheet(wb: Workbook, sheet_name: str, filename: str) -> Path:
    SPLIT_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    src_ws = wb[sheet_name]
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = sheet_name
    for row in src_ws.iter_rows():
        for cell in row:
            out_cell = out_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                out_cell.font = copy(cell.font)
                out_cell.fill = copy(cell.fill)
                out_cell.border = copy(cell.border)
                out_cell.alignment = copy(cell.alignment)
                out_cell.number_format = cell.number_format
                out_cell.protection = copy(cell.protection)
    for key, dimension in src_ws.column_dimensions.items():
        out_ws.column_dimensions[key].width = dimension.width
    out_ws.freeze_panes = src_ws.freeze_panes
    out_path = SPLIT_OUTPUT_DIR / filename
    rows = [list(row) for row in out_ws.iter_rows(min_row=2, values_only=True)]
    headers = [cell.value for cell in out_ws[1]]
    write_standard_json(out_path, headers, rows)
    sync_table(f"meituan_{Path(filename).stem}", headers, rows)
    try:
        out_wb.save(out_path)
    except PermissionError:
        print(f"WARN Excel file is open, skipped writing: {out_path}")
    return out_path


def save_to_excel(overview_rows: list[list[Any]], ranking_rows: list[list[Any]], output_path=DEFAULT_EXCEL_PATH) -> Path:
    overview_headers = [*OVERVIEW_HEADERS, "hotel_id"]
    ranking_headers = [*RANKING_HEADERS, "hotel_id"]
    overview_rows = [list(row) + [HOTEL_ID] for row in overview_rows]
    ranking_rows = [list(row) + [HOTEL_ID] for row in ranking_rows]
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = get_or_create_workbook(output_path)
    overview_ws = reset_sheet(wb, OVERVIEW_SHEET, overview_headers, [20, 12, 14, 14, 16, 16, 16, 16, 18, 18, 16, 16])
    for row in overview_rows:
        overview_ws.append(row)
        overview_ws.cell(overview_ws.max_row, 1).number_format = "yyyy-mm-dd hh:mm:ss"
    ranking_ws = reset_sheet(wb, RANKING_SHEET, ranking_headers, [20, 12, 20, 12, 34, 16, 16])
    for row in ranking_rows:
        ranking_ws.append(row)
        ranking_ws.cell(ranking_ws.max_row, 1).number_format = "yyyy-mm-dd hh:mm:ss"
    wb.save(output_path)
    save_single_sheet(wb, OVERVIEW_SHEET, "ota_review_overview.xlsx")
    save_single_sheet(wb, RANKING_SHEET, "ota_review_ranking.xlsx")
    return output_path


def sample_payload() -> dict[str, Any]:
    return {
        "score": 47,
        "subScores": [
            {"title": "\u4f4d\u7f6e", "score": 4.7},
            {"title": "\u8bbe\u65bd", "score": 4.7},
            {"title": "\u670d\u52a1", "score": 4.8},
            {"title": "\u536b\u751f", "score": 4.7},
        ],
        "tagPos": [{"tagName": "\u670d\u52a1\u70ed\u60c5", "count": 583, "type": 1}],
        "commentAnalysisInfo": {"totalCount365": 834, "badPercent365": 0.0408},
    }


def collect(
    payload: dict[str, Any] | None = None,
    ranking_payload: dict[str, Any] | None = None,
    dianping_payload: dict[str, Any] | None = None,
) -> tuple[list[list[Any]], list[list[Any]]]:
    captured_at = datetime.now()
    payload = payload or MeituanReviewClient(MEITUAN_ME_COOKIE).get_contrast_new()
    overview_rows = normalize_overview_rows(payload, captured_at)
    if dianping_payload:
        overview_rows.extend(normalize_overview_rows(dianping_payload, captured_at, "dianping"))
    ranking_rows = normalize_ranking_rows(payload, captured_at)
    if ranking_payload:
        ranking_rows.extend(normalize_keyword_ranking_rows(ranking_payload, captured_at))
    return overview_rows, ranking_rows


def main() -> None:
    parser = argparse.ArgumentParser(description="Meituan review data crawler.")
    parser.add_argument("--self-test", action="store_true", help="Parse built-in sample data; no request.")
    parser.add_argument("--contrast-url", default=CONTRAST_URL, help="Signed contrast URL copied from Network.")
    parser.add_argument("--ranking-url", default=RANKING_URL, help="Review keyword ranking URL.")
    args = parser.parse_args()
    if args.self_test:
        overview_rows, ranking_rows = collect(sample_payload())
    else:
        if not MEITUAN_ME_COOKIE:
            raise RuntimeError("Please set MEITUAN_ME_COOKIE or MEITUAN_COOKIE in meituan_config.py")
        client = MeituanReviewClient(MEITUAN_ME_COOKIE, args.contrast_url)
        dianping_payload = client.get_dianping_contrast() if DIANPING_CONTRAST_URL else None
        overview_rows, ranking_rows = collect(
            client.get_contrast_new(), client.get_ranking(args.ranking_url), dianping_payload
        )
    excel_path = save_to_excel(overview_rows, ranking_rows)
    print(f"review_overview rows={len(overview_rows)}")
    print(f"review_ranking rows={len(ranking_rows)}")
    print(f"Excel saved: {excel_path}")


if __name__ == "__main__":
    main()
