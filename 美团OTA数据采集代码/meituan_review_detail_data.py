from __future__ import annotations

import argparse
import json
import os
import sys
import time
from datetime import date, datetime, timedelta, timezone
from http.cookies import SimpleCookie
from pathlib import Path
from typing import Any
from urllib.parse import urlencode

import pymysql
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from playwright.sync_api import sync_playwright

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from meituan_config import BIZ_ACCOUNT_ID, HOTEL_NAME, MEITUAN_ME_COOKIE, PARTNER_ID, POI_ID, USER_AGENT
from meituan_review_overview_sync import extract_overview_counts, sync_overview_counts
from ota_mysql_writer import DB_CONFIG, OUTPUT_DIR


API_URL = os.environ.get(
    "MEITUAN_REVIEW_DETAIL_URL",
    "https://me.meituan.com/api/gw/v1/base/comments/queryGeneralCommentInfo",
).strip()
DIANPING_API_URL = os.environ.get("MEITUAN_DIANPING_REVIEW_DETAIL_URL", "").strip()
OUTPUT_PATH = OUTPUT_DIR / "ota_review_detail.xlsx"
TABLE_NAME = "meituan_ota_review_detail"
HOTEL_ID = os.environ.get("HOTEL_ID", "").strip()
HEADERS = [
    "snapshot_time",
    "channel_source",
    "hotel_name",
    "poi_id",
    "review_id",
    "reviewer_name_masked",
    "review_score",
    "review_content",
    "review_time",
    "stay_date",
    "merchant_reply_content",
    "merchant_reply_time",
    "is_replied",
    "room_type_name",
    "ota_product_name",
    "has_image",
    "image_count",
    "image_urls_json",
    "is_anonymous",
    "is_negative_review",
    "read_status",
    "hygiene_score",
    "facility_score",
    "location_score",
    "service_score",
]
CHINA_TZ = timezone(timedelta(hours=8))
COMMENT_PAGE_URL = "https://me.meituan.com/ebooking/merchant/comment-manage-react"


class MeituanReviewDetailClient:
    def __init__(self, cookie: str, api_url: str = API_URL, platform: int = 1):
        self.endpoint = api_url.split("?", 1)[0]
        self.platform = platform
        self.playwright = sync_playwright().start()
        self.browser = self.playwright.chromium.launch(headless=True)
        self.context = self.browser.new_context(user_agent=USER_AGENT, viewport={"width": 1440, "height": 1000})
        parsed = SimpleCookie()
        parsed.load(cookie)
        cookies = [
            {"name": name, "value": morsel.value, "domain": ".meituan.com", "path": "/"}
            for name, morsel in parsed.items()
        ]
        if not cookies:
            self.close()
            raise RuntimeError("MEITUAN_ME_COOKIE is empty or invalid")
        self.context.add_cookies(cookies)
        self.page = self.context.new_page()
        try:
            with self.page.expect_response(
                lambda response: self.endpoint in response.url and "replyType=0" in response.url,
                timeout=45000,
            ):
                self.page.goto(COMMENT_PAGE_URL, wait_until="domcontentloaded", timeout=60000)
        except Exception:
            self.close()
            raise

    def fetch_page(self, offset: int, limit: int) -> dict[str, Any]:
        query = urlencode(
            {
                "poiId": POI_ID,
                "partnerId": PARTNER_ID,
                "platform": self.platform,
                "tag": "",
                "keywords": "",
                "replyType": 0,
                "offset": offset,
                "limit": limit,
                "bizAccountId": BIZ_ACCOUNT_ID,
                "yodaReady": "h5",
                "csecplatform": 4,
                "csecversion": "4.2.4",
            }
        )
        payload = self.page.evaluate(
            """async (url) => {
                const response = await fetch(url, {
                    credentials: "include",
                    headers: {"X-Requested-With": "XMLHttpRequest", "Request-Page-Source": "ME"}
                });
                if (!response.ok) throw new Error(`HTTP ${response.status}`);
                return await response.json();
            }""",
            f"{self.endpoint}?{query}",
        )
        if payload.get("code") != 10000:
            raise RuntimeError(f"Review detail request failed: code={payload.get('code')}")
        return payload

    def close(self) -> None:
        for resource in ("context", "browser", "playwright"):
            item = getattr(self, resource, None)
            if item:
                try:
                    item.close() if resource != "playwright" else item.stop()
                except Exception:
                    pass
                setattr(self, resource, None)


def millis_to_datetime(value: Any) -> datetime | None:
    if value in (None, ""):
        return None
    return datetime.fromtimestamp(float(value) / 1000, CHINA_TZ).replace(tzinfo=None)


def millis_to_date(value: Any) -> date | None:
    parsed = millis_to_datetime(value)
    return parsed.date() if parsed else None


def normalize_score(value: Any) -> float | None:
    if value in (None, ""):
        return None
    number = float(value)
    return round(number / 10 if number > 5 else number, 1)


def score_map(items: Any) -> dict[str, float | None]:
    result: dict[str, float | None] = {}
    for item in items or []:
        result[str(item.get("title") or "")] = normalize_score(item.get("score"))
    return result


def normalize_rows(payload: dict[str, Any], captured_at: datetime | None = None) -> list[list[Any]]:
    captured_at = captured_at or datetime.now()
    rows: list[list[Any]] = []
    for item in payload.get("data", {}).get("commentList") or []:
        review_id = str(item.get("id") or "").strip()
        if not review_id:
            continue
        images = item.get("imgUrls") or []
        scores = score_map(item.get("scoreList"))
        reply = str(item.get("bizReply") or "").strip()
        is_replied = bool(item.get("replyId") or reply)
        hotel_name = str(item.get("poiName") or HOTEL_NAME).strip()
        rows.append(
            [
                captured_at,
                "美团",
                hotel_name,
                POI_ID,
                review_id,
                str(item.get("userName") or "").strip(),
                normalize_score(item.get("accurateScore") or item.get("score")),
                str(item.get("comment") or "").strip(),
                millis_to_datetime(item.get("commentTime")),
                millis_to_date(item.get("consumeTime")),
                reply,
                millis_to_datetime(item.get("replyTime")) if is_replied else None,
                is_replied,
                str(item.get("roomName") or "").strip(),
                str(item.get("dealName") or "").strip(),
                bool(images),
                len(images),
                json.dumps(images, ensure_ascii=False),
                bool(item.get("anonymous")),
                bool(item.get("badComment")),
                item.get("readStatus"),
                scores.get("卫生"),
                scores.get("设施"),
                scores.get("位置"),
                scores.get("服务"),
            ]
        )
    return rows


def collect_online(
    limit: int,
    max_pages: int,
    full_history: bool,
    existing_ids: set[str] | None = None,
) -> tuple[list[list[Any]], dict[str, int]]:
    client = MeituanReviewDetailClient(MEITUAN_ME_COOKIE)
    existing_ids = existing_ids or set()
    captured_at = datetime.now()
    rows_by_id: dict[str, list[Any]] = {}
    offset = 1
    total = None
    overview_counts: dict[str, int] = {}
    previous_ids: set[str] = set()
    try:
        while full_history or offset <= max_pages:
            if offset > 1 and (offset - 1) % 40 == 0:
                print(f"review_detail browser restart before page={offset}")
                client.close()
                time.sleep(1)
                client = MeituanReviewDetailClient(MEITUAN_ME_COOKIE)
            for attempt in range(1, 4):
                try:
                    payload = client.fetch_page(offset, limit)
                    break
                except Exception as exc:
                    client.close()
                    if attempt == 3:
                        raise
                    print(f"review_detail page={offset} retry={attempt}/3 error={exc}")
                    time.sleep(attempt * 2)
                    client = MeituanReviewDetailClient(MEITUAN_ME_COOKIE)
            data = payload.get("data") or {}
            if offset == 1:
                overview_counts = extract_overview_counts(payload)
            total = int(data.get("total") or 0)
            page_rows = normalize_rows(payload, captured_at)
            page_ids = {str(row[4]) for row in page_rows}
            print(f"review_detail page={offset} rows={len(page_rows)} total={total}")
            if not page_rows or page_ids == previous_ids:
                break
            for row in page_rows:
                rows_by_id[str(row[4])] = row
            if not full_history and page_ids & existing_ids:
                print(f"review_detail incremental stop: page={offset} contains existing review_id")
                break
            if total and len(rows_by_id) >= total:
                break
            previous_ids = page_ids
            offset += 1
            time.sleep(0.25)
    finally:
        client.close()
    rows = sorted(rows_by_id.values(), key=lambda row: row[8] or datetime.min, reverse=True)
    return rows, overview_counts


def collect_overview_counts(api_url: str, platform: int) -> dict[str, int]:
    if not api_url:
        print(f"review_detail platform={platform} overview sync skipped: URL is empty")
        return {}
    client = MeituanReviewDetailClient(MEITUAN_ME_COOKIE, api_url, platform)
    try:
        return extract_overview_counts(client.fetch_page(1, 10))
    finally:
        client.close()


def load_existing_review_ids() -> set[str]:
    connection = pymysql.connect(**DB_CONFIG)
    try:
        with connection.cursor() as cursor:
            if HOTEL_ID:
                cursor.execute(f"SELECT review_id FROM {TABLE_NAME} WHERE hotel_id=%s", (HOTEL_ID,))
            else:
                cursor.execute(f"SELECT review_id FROM {TABLE_NAME}")
            return {str(row[0]) for row in cursor.fetchall() if row[0] is not None}
    finally:
        connection.close()


def json_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    return value


def save_outputs(headers: list[str], rows: list[list[Any]]) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "评价明细"
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    for index, width in enumerate([20, 10, 28, 14, 16, 16, 12, 50, 20, 20, 50, 20, 12, 35, 35, 12, 12, 35, 12, 14, 12, 12, 12, 12, 12], 1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width
    workbook.save(OUTPUT_PATH)

    payload = {
        "table_name": OUTPUT_PATH.stem,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "row_count": len(rows),
        "rows": [{header: json_value(row[index]) for index, header in enumerate(headers)} for row in rows],
    }
    OUTPUT_PATH.with_suffix(".json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def upsert_mysql(headers: list[str], rows: list[list[Any]]) -> None:
    if not rows:
        print("DB sync skipped: no review detail rows")
        return

    columns = ", ".join(f"`{name}`" for name in headers)
    placeholders = ", ".join(["%s"] * len(headers))
    updates = ", ".join(
        f"`{name}`=VALUES(`{name}`)"
        for name in headers
        if name not in {"channel_source", "poi_id", "review_id"}
    )
    sql = f"INSERT INTO `{TABLE_NAME}` ({columns}) VALUES ({placeholders}) ON DUPLICATE KEY UPDATE {updates}"
    connection = pymysql.connect(**DB_CONFIG, autocommit=False)
    try:
        with connection.cursor() as cursor:
            cursor.executemany(sql, rows)
        connection.commit()
    except Exception:
        connection.rollback()
        raise
    finally:
        connection.close()
    print(f"DB upserted: {TABLE_NAME} rows={len(rows)}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Collect Meituan review details.")
    parser.add_argument("--input-json", help="Parse a saved response without requesting.")
    parser.add_argument("--limit", type=int, default=10)
    parser.add_argument("--max-pages", type=int, default=500)
    parser.add_argument("--full-history", action="store_true")
    parser.add_argument("--no-db", action="store_true")
    args = parser.parse_args()

    if args.input_json:
        payload = json.loads(Path(args.input_json).read_text(encoding="utf-8-sig"))
        rows = normalize_rows(payload)
        overview_counts = extract_overview_counts(payload)
    else:
        if not MEITUAN_ME_COOKIE:
            raise RuntimeError("MEITUAN_ME_COOKIE is empty")
        existing_ids = set() if args.no_db else load_existing_review_ids()
        full_history = args.full_history or (not args.no_db and not existing_ids)
        if full_history:
            print("review_detail mode=full_history")
        else:
            print(f"review_detail mode=incremental existing={len(existing_ids)}")
        rows, overview_counts = collect_online(
            max(1, args.limit), max(1, args.max_pages), full_history, existing_ids
        )
    dianping_counts = collect_overview_counts(DIANPING_API_URL, 0) if not args.input_json else {}
    headers = [*HEADERS, "hotel_id"]
    rows = [list(row) + [HOTEL_ID] for row in rows]
    save_outputs(headers, rows)
    if not args.no_db:
        upsert_mysql(headers, rows)
        sync_overview_counts(overview_counts, HOTEL_ID)
        sync_overview_counts(dianping_counts, HOTEL_ID, "dianping")
    print(f"review_detail rows={len(rows)}")
    print(f"Excel saved: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
