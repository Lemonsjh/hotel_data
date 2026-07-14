from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from meituan_config import MEITUAN_ME_COOKIE, PARTNER_ID, POI_ID, USER_AGENT

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "OTA采集服务"))
from mapping_product_sync import sync_meituan_products
from ota_mysql_writer import DB_CONFIG, OUTPUT_DIR, sync_table


TABLE_NAME = "ota_goods_price_mapping"
HOTEL_ID = os.environ.get("HOTEL_ID", "").strip()
PLATFORM = "\u7f8e\u56e2"

QUERY_URL = os.environ.get("MEITUAN_GOODS_QUERY_URL", "").strip()
CALC_PRICE_URL = os.environ.get("MEITUAN_CALC_PRICE_URL", "").strip()
PRICE_STATUS_URL = os.environ.get("MEITUAN_PRICE_STATUS_URL", "").strip()
PRICE_STATUS_PAYLOAD_PATH = os.environ.get("MEITUAN_PRICE_STATUS_PAYLOAD_FILE", "").strip()
DEFAULT_PRICE_STATUS_PAYLOAD_FILE = Path(PRICE_STATUS_PAYLOAD_PATH) if PRICE_STATUS_PAYLOAD_PATH else None

HEADERS = [
    "snapshot_time",
    "channel_source",
    "ota_room_type_id",
    "room_type_name",
    "business_date",
    "ota_product_id",
    "ota_product_name",
    "rate_plan_name",
    "is_super_deal",
    "ota_sale_price",
    "commission_rate",
]
PRODUCT_ID_INDEX = HEADERS.index("ota_product_id")
SUPER_DEAL_INDEX = HEADERS.index("is_super_deal")


class MeituanGoodsClient:
    def __init__(self, cookie: str, query_url: str = QUERY_URL):
        self.query_url = query_url
        self.session = requests.Session()
        self.session.headers.update(
            {
                "User-Agent": USER_AGENT,
                "Accept": "application/json, text/plain, */*",
                "Referer": "https://me.meituan.com/ebooking/merchant/price",
                "Origin": "https://me.meituan.com",
                "X-Requested-With": "XMLHttpRequest",
                "Request-Page-Source": "ME",
            }
        )
        if cookie.strip():
            self.session.headers["Cookie"] = cookie.strip()

    def query_goods(self) -> dict[str, Any]:
        if not self.query_url:
            raise RuntimeError("MEITUAN_GOODS_QUERY_URL is empty; configure a current signed URL")
        response = self.session.post(
            self.query_url,
            json={"poiId": str(POI_ID), "partnerId": int(PARTNER_ID)},
            timeout=30,
        )
        response.raise_for_status()
        payload = response.json()
        if payload.get("code") not in (0, 10000, "0", "10000"):
            raise RuntimeError({"code": payload.get("code"), "error": payload.get("error")})
        data = payload.get("data")
        if not isinstance(data, dict):
            raise RuntimeError("queryListAndTag data is not object")
        return data

    def query_calc_price(self, goods_data: dict[str, Any], calc_price_url: str = CALC_PRICE_URL) -> dict[str, Any]:
        if not calc_price_url:
            raise RuntimeError("MEITUAN_CALC_PRICE_URL is empty; configure a current signed URL")
        data = self._request_calc_price(calc_price_url, build_calc_price_payload(goods_data))
        if data is None:
            raise RuntimeError("calcPriceV2 failed")
        return data

    def _request_calc_price(self, calc_price_url: str, payload: dict[str, Any]) -> dict[str, Any] | None:
        response = self.session.post(calc_price_url, json=payload, timeout=30)
        response.raise_for_status()
        data = response.json()
        if data.get("code") in (0, 10000, "0", "10000") and isinstance(data.get("data"), dict):
            return data["data"]
        return None

    def query_price_status(self, goods_data: dict[str, Any], price_status_url: str, business_date: str) -> list[Any]:
        if not price_status_url:
            raise RuntimeError("MEITUAN_PRICE_STATUS_URL is empty; configure a current signed URL")
        response = self.session.post(price_status_url, json=build_price_status_payload(goods_data, business_date), timeout=30)
        response.raise_for_status()
        payload = response.json()
        if payload.get("code") not in (0, 10000, "0", "10000"):
            raise RuntimeError({"code": payload.get("code"), "error": payload.get("error")})
        data = payload.get("data")
        if not isinstance(data, list):
            raise RuntimeError("queryPriceInventoryStatusInfo data is not list")
        return data


def to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def compact_json(value: Any) -> str:
    if value in (None, "", [], {}):
        return ""
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"))


def cents_to_yuan(value: Any) -> Any:
    if value in (None, ""):
        return ""
    try:
        return round(float(value) / 100, 2)
    except (TypeError, ValueError):
        return value


def ratio_to_percent(value: Any) -> str:
    if value in (None, ""):
        return ""
    try:
        return f"{float(value) / 100:.2f}%"
    except (TypeError, ValueError):
        return str(value)


def visible_tag_names(goods: dict[str, Any]) -> str:
    tags = ((goods.get("tagInfo") or {}).get("goodsTags") or [])
    names = [to_text(item.get("tagName")) for item in tags if isinstance(item, dict) and item.get("visible")]
    return ",".join(name for name in names if name)


def iter_goods(data: dict[str, Any]) -> list[dict[str, Any]]:
    goods_rows: list[dict[str, Any]] = []
    for real_room in data.get("realRoomRelations") or []:
        if not isinstance(real_room, dict):
            continue
        for logic_room in real_room.get("logicRoomRelations") or []:
            if not isinstance(logic_room, dict):
                continue
            for goods in logic_room.get("goodsList") or []:
                if isinstance(goods, dict):
                    goods_rows.append(goods)
    return goods_rows


def iter_room_ids(data: dict[str, Any]) -> list[Any]:
    room_ids: list[Any] = []
    for real_room in data.get("realRoomRelations") or []:
        for logic_room in real_room.get("logicRoomRelations") or [] if isinstance(real_room, dict) else []:
            room = logic_room.get("roomBaseInfo") or {}
            room_id = room.get("roomId")
            if room_id and room_id not in room_ids:
                room_ids.append(room_id)
    return room_ids


def build_price_status_payload(goods_data: dict[str, Any], business_date: str) -> dict[str, Any]:
    return {
        "startDate": business_date,
        "endDate": business_date,
        "poiId": str(POI_ID),
        "partnerId": int(PARTNER_ID),
        "goodsIds": [goods.get("goodsId") for goods in iter_goods(goods_data) if goods.get("goodsId")],
        "roomIds": iter_room_ids(goods_data),
    }


def build_calc_price_payload(goods_data: dict[str, Any]) -> dict[str, Any]:
    goods_list = []
    for goods in iter_goods(goods_data):
        goods_list.append(
            {
                "goodsBaseInfo": goods,
                "operateType": 6,
                "priceRecordWay": 8,
                "weekDiff": False,
                "ratioConfig": {
                    "ratioChange": False,
                    "newRatio": "",
                    "ratioType": 2,
                },
                "calcPriceUnifiedDateModel": {
                    "dates": [
                        {
                            "startDate": datetime.now().strftime("%Y-%m-%d"),
                            "endDate": datetime.now().strftime("%Y-%m-%d"),
                        }
                    ],
                    "calcPriceWeekModels": [
                        {
                            "inWeek": [1, 2, 3, 4, 5, 6, 7],
                            "calcPriceInfo": {
                                "salePrice": {"operateType": 3, "operateNum": ""},
                                "basePrice": {"operateType": 3, "operateNum": ""},
                                "subPrice": {"operateType": 3, "operateNum": ""},
                            },
                            "calcPriceFactorInfos": [],
                        }
                    ],
                },
            }
        )
    return {
        "currency": "CNY",
        "partnerId": int(PARTNER_ID),
        "poiId": str(POI_ID),
        "goodsList": goods_list,
    }


def load_payload_file(path: Path) -> dict[str, Any]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise RuntimeError("payload file JSON is not object")
    return payload


def payload_business_date(payload: dict[str, Any], default: str) -> str:
    return str(payload.get("startDate") or default)


def is_super_deal(goods: dict[str, Any]) -> bool:
    rp_name = to_text(goods.get("rpCustomName"))
    activity_map = goods.get("goodsActivityMap") or {}
    tag_names = visible_tag_names(goods)
    return (
        "\u8d85\u7ea7\u56e2\u8d2d" in rp_name
        or "\u8d85\u56e2" in tag_names
        or activity_map.get("superDeal") == 1
        or goods.get("superDealReSale") is True
    )


def block_reason(goods: dict[str, Any]) -> str:
    if is_super_deal(goods):
        return "super_deal"
    if goods.get("canAdjustPrice") is False:
        return "can_adjust_price_false"
    return ""


def first_price_info(detail: dict[str, Any]) -> tuple[str, dict[str, Any]]:
    for real_info in detail.get("realPriceInfos") or []:
        for week_info in real_info.get("weekPriceInfos") or []:
            price = week_info.get("originalPriceInfo") or week_info.get("priceInfo") or {}
            if price:
                return real_info.get("startDate") or "", price
    unified = detail.get("unifiedDatePriceInfos") or {}
    date = ((unified.get("dates") or [{}])[0] or {}).get("startDate") or ""
    week_info = (unified.get("weekPriceInfos") or [{}])[0] or {}
    return date, week_info.get("originalPriceInfo") or week_info.get("priceInfo") or {}


def calc_price_index(price_data: dict[str, Any]) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    for detail in price_data.get("goodsDetails") or []:
        if not isinstance(detail, dict):
            continue
        goods = detail.get("goodsBaseInfo") or {}
        ota_product_id = goods.get("goodsId")
        date, price = first_price_info(detail)
        if ota_product_id:
            result[str(ota_product_id)] = {
                "business_date": date,
                "ota_sale_price": cents_to_yuan(price.get("salePrice")),
                "commission_rate": ratio_to_percent(price.get("subRatio")),
                "goods": goods,
            }
    return result


def price_status_index(price_status_data: list[Any], business_date: str) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    for item in price_status_data:
        if not isinstance(item, dict):
            continue
        goods = item.get("goodsBaseInfo") or {}
        ota_product_id = goods.get("goodsId")
        price_items = ((item.get("goodsPriceMap") or {}).get(business_date) or [])
        price = price_items[0] if price_items and isinstance(price_items[0], dict) else {}
        if ota_product_id and price:
            result[str(ota_product_id)] = {
                "business_date": price.get("date") or business_date,
                "ota_sale_price": cents_to_yuan(price.get("salePrice")),
                "commission_rate": ratio_to_percent(price.get("subRatio")),
                "goods": goods,
            }
    return result


def payload_goods_index(payload: dict[str, Any]) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    for item in payload.get("goodsList") or []:
        if not isinstance(item, dict):
            continue
        goods = item.get("goodsBaseInfo") or {}
        ota_product_id = goods.get("goodsId")
        if ota_product_id:
            result[str(ota_product_id)] = goods
    return result


def normalize_rows(
    data: dict[str, Any],
    captured_at: datetime,
    price_data: dict[str, Any] | list[Any] | None = None,
    business_date: str = "",
) -> list[list[Any]]:
    rows: list[list[Any]] = []
    prices = price_status_index(price_data, business_date) if isinstance(price_data, list) else calc_price_index(price_data or {})
    for real_room in data.get("realRoomRelations") or []:
        if not isinstance(real_room, dict):
            continue
        for logic_room in real_room.get("logicRoomRelations") or []:
            if not isinstance(logic_room, dict):
                continue
            room = logic_room.get("roomBaseInfo") or {}
            for goods in logic_room.get("goodsList") or []:
                if not isinstance(goods, dict):
                    continue
                price = prices.get(str(goods.get("goodsId")), {})
                rows.append(
                    [
                        captured_at,
                        PLATFORM,
                        room.get("roomId"),
                        room.get("roomName"),
                        price.get("business_date", ""),
                        goods.get("goodsId"),
                        goods.get("goodsName"),
                        goods.get("rpCustomName"),
                        is_super_deal(goods),
                        price.get("ota_sale_price", ""),
                        price.get("commission_rate", ""),
                    ]
                )
    if rows or not prices:
        return rows
    for ota_product_id, price in prices.items():
        goods = price.get("goods") or {}
        rows.append(
            [
                captured_at,
                PLATFORM,
                "",
                "",
                price.get("business_date", ""),
                ota_product_id,
                goods.get("goodsName"),
                goods.get("rpCustomName"),
                is_super_deal(goods),
                price.get("ota_sale_price", ""),
                price.get("commission_rate", ""),
            ]
        )
    return rows


def filter_rows_by_goods_ids(rows: list[list[Any]], goods_ids: set[str]) -> list[list[Any]]:
    if not goods_ids:
        return rows
    return [row for row in rows if str(row[PRODUCT_ID_INDEX]) in goods_ids]


def json_safe(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return value


def write_json(path: Path, headers: list[str], rows: list[list[Any]]) -> Path:
    json_path = path.with_suffix(".json")
    payload = {
        "table_name": path.stem,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "row_count": len(rows),
        "rows": [
            {header: json_safe(row[index]) if index < len(row) else "" for index, header in enumerate(headers)}
            for row in rows
        ],
    }
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return json_path


def sync_room_type_mapping() -> None:
    try:
        import pymysql
    except ImportError:
        return
    connection = pymysql.connect(**DB_CONFIG, autocommit=True)
    try:
        with connection.cursor() as cursor:
            stats = sync_meituan_products(cursor)
        print("DB synced: hotel_room_type_mapping")
        print(f"Mapping products refreshed: {stats}")
    finally:
        connection.close()


def save_excel(rows: list[list[Any]]) -> Path:
    headers = [*HEADERS, "hotel_id"]
    rows = [list(row) + [HOTEL_ID] for row in rows]
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUTPUT_DIR / f"{TABLE_NAME}.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = TABLE_NAME
    ws.append(headers)
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append(row)
        ws.cell(ws.max_row, 1).number_format = "yyyy-mm-dd hh:mm:ss"
    widths = [20, 12, 16, 26, 16, 16, 46, 28, 14, 18, 22, 16]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=index).column_letter].width = width
    wb.save(out_path)
    write_json(out_path, headers, rows)
    sync_table(f"meituan_{TABLE_NAME}", headers, rows)
    sync_room_type_mapping()
    return out_path


def load_sample_file(path: Path) -> dict[str, Any]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    data = payload.get("data") if isinstance(payload, dict) else None
    if not isinstance(data, dict):
        raise RuntimeError("sample file data is not object")
    return data


def collect(
    query_url: str = QUERY_URL,
    sample_file: Path | None = None,
    price_url: str = CALC_PRICE_URL,
    price_sample_file: Path | None = None,
    payload_file: Path | None = None,
    price_status_payload_file: Path | None = None,
    price_status_url: str = PRICE_STATUS_URL,
    business_date: str | None = None,
) -> list[list[Any]]:
    client = MeituanGoodsClient(MEITUAN_ME_COOKIE, query_url)
    business_date = business_date or datetime.now().strftime("%Y-%m-%d")
    if sample_file:
        data = load_sample_file(sample_file)
    else:
        data = client.query_goods()
    payload_goods_ids: set[str] = set()
    if price_sample_file:
        price_data = load_sample_file(price_sample_file)
    elif price_status_payload_file:
        if not price_status_url:
            raise RuntimeError("MEITUAN_PRICE_STATUS_URL is empty; configure a current signed URL")
        payload = load_payload_file(price_status_payload_file)
        payload_goods_ids = {str(item) for item in payload.get("goodsIds") or []}
        business_date = payload_business_date(payload, business_date)
        response = client.session.post(price_status_url, json=payload, timeout=30)
        response.raise_for_status()
        result = response.json()
        if result.get("code") not in (0, 10000, "0", "10000"):
            raise RuntimeError({"code": result.get("code"), "error": result.get("error")})
        price_data = result.get("data")
        if not isinstance(price_data, list):
            raise RuntimeError("queryPriceInventoryStatusInfo data is not list")
    elif payload_file:
        if not price_url:
            raise RuntimeError("MEITUAN_CALC_PRICE_URL is empty; configure a current signed URL")
        payload = load_payload_file(payload_file)
        payload_goods_ids = set(payload_goods_index(payload))
        price_data = client._request_calc_price(price_url, payload)
        if price_data is None:
            raise RuntimeError("calcPriceV2 failed with payload file")
    else:
        price_data = client.query_price_status(data, price_status_url, business_date)
    return filter_rows_by_goods_ids(normalize_rows(data, datetime.now(), price_data, business_date), payload_goods_ids)


def main() -> None:
    parser = argparse.ArgumentParser(description="Meituan goods price mapping crawler.")
    parser.add_argument("--url", default=QUERY_URL, help="Signed queryListAndTag URL copied from Network.")
    parser.add_argument("--price-url", default=CALC_PRICE_URL, help="Signed calcPriceV2 URL copied from Network.")
    parser.add_argument("--price-status-url", default=PRICE_STATUS_URL, help="Signed queryPriceInventoryStatusInfo URL.")
    parser.add_argument("--business-date", help="Business date, YYYY-MM-DD. Default is today.")
    parser.add_argument("--sample-file", type=Path, help="Parse saved response JSON; no request.")
    parser.add_argument("--price-sample-file", type=Path, help="Parse saved calcPriceV2 response JSON.")
    parser.add_argument("--payload-file", type=Path, help="Use saved calcPriceV2 request payload JSON.")
    parser.add_argument(
        "--price-status-payload-file",
        type=Path,
        default=(
            DEFAULT_PRICE_STATUS_PAYLOAD_FILE
            if DEFAULT_PRICE_STATUS_PAYLOAD_FILE and DEFAULT_PRICE_STATUS_PAYLOAD_FILE.exists()
            else None
        ),
        help="Use saved queryPriceInventoryStatusInfo request payload JSON.",
    )
    args = parser.parse_args()
    if not args.sample_file and not MEITUAN_ME_COOKIE:
        raise RuntimeError("Please set MEITUAN_ME_COOKIE or MEITUAN_COOKIE in meituan_config.py")
    rows = collect(
        args.url,
        args.sample_file,
        args.price_url,
        args.price_sample_file,
        args.payload_file,
        args.price_status_payload_file,
        args.price_status_url,
        args.business_date,
    )
    out_path = save_excel(rows)
    super_deal = sum(1 for row in rows if row[SUPER_DEAL_INDEX] is True)
    print(f"goods rows={len(rows)} super_deal={super_deal} output={out_path}")


if __name__ == "__main__":
    main()
