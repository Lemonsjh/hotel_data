from __future__ import annotations

import argparse
import json
import os
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any
from urllib.parse import urlencode

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from meituan_config import HOTEL_NAME, MEITUAN_EB_COOKIE, POI_ID, USER_AGENT

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from ota_mysql_writer import OUTPUT_DIR, sync_table


OUTPUT_PATH = OUTPUT_DIR / "ota_nearby_event.xlsx"
TABLE_NAME = "meituan_ota_nearby_event"
HOTEL_ID = os.environ.get("HOTEL_ID", "").strip()
HEADERS = [
    "snapshot_time",
    "channel_source",
    "hotel_name",
    "poi_id",
    "event_id",
    "event_class_id",
    "event_name",
    "event_start_date",
    "event_end_date",
    "event_address",
    "distance_km",
    "countdown_days",
]


def build_url() -> str:
    query = {
        "poiId": POI_ID,
        "yodaReady": "h5",
        "csecplatform": 4,
        "csecversion": "4.2.4",
    }
    return f"https://eb.meituan.com/api/shepherdGw/dft/event/comingEventInfo?{urlencode(query)}"


class MeituanNearbyEventClient:
    def __init__(self, cookie: str, url: str):
        self.url = url
        self.session = requests.Session()
        self.session.headers.update(
            {
                "User-Agent": USER_AGENT,
                "Accept": "application/json, text/plain, */*",
                "Referer": "https://eb.meituan.com/",
            }
        )
        if cookie.strip():
            self.session.headers["Cookie"] = cookie.strip()

    def fetch(self) -> dict[str, Any]:
        response = self.session.get(self.url, timeout=30)
        response.raise_for_status()
        payload = response.json()
        if payload.get("status") != 0:
            raise RuntimeError(f"Nearby event request failed: status={payload.get('status')}")
        return payload


def parse_date(value: Any) -> date | None:
    if not value:
        return None
    return datetime.strptime(str(value), "%Y-%m-%d").date()


def normalize_rows(payload: dict[str, Any], snapshot_time: datetime | None = None) -> list[list[Any]]:
    snapshot_time = snapshot_time or datetime.now()
    items = payload.get("data", {}).get("list") or []
    rows: list[list[Any]] = []
    seen_event_ids: set[str] = set()
    for item in items:
        event_id = str(item.get("eventId") or "").strip()
        if not event_id or event_id in seen_event_ids:
            continue
        seen_event_ids.add(event_id)
        rows.append(
            [
                snapshot_time,
                "美团",
                HOTEL_NAME,
                POI_ID,
                event_id,
                item.get("eventClassId"),
                str(item.get("name") or "").strip(),
                parse_date(item.get("startDate")),
                parse_date(item.get("endDate")),
                str(item.get("address") or "").strip(),
                item.get("distance"),
                item.get("countdownDate"),
            ]
        )
    return sorted(rows, key=lambda row: (row[7] or date.max, row[10] or 0, row[4]))


def json_value(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    return value


def save_outputs(rows: list[list[Any]], sync_db: bool = True) -> None:
    headers = [*HEADERS, "hotel_id"]
    rows = [list(row) + [HOTEL_ID] for row in rows]
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "周边事件"
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    widths = [20, 12, 28, 15, 15, 16, 45, 16, 16, 35, 14, 14]
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = width
    for cells in sheet.iter_rows(min_row=2):
        cells[0].number_format = "yyyy-mm-dd hh:mm:ss"
        cells[7].number_format = "yyyy-mm-dd"
        cells[8].number_format = "yyyy-mm-dd"
    workbook.save(OUTPUT_PATH)

    payload = {
        "table_name": OUTPUT_PATH.stem,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "row_count": len(rows),
        "rows": [
            {header: json_value(row[index]) for index, header in enumerate(headers)}
            for row in rows
        ],
    }
    OUTPUT_PATH.with_suffix(".json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    if sync_db:
        sync_table(TABLE_NAME, headers, rows)


def sample_payload() -> dict[str, Any]:
    return {
        "status": 0,
        "data": {
            "list": [
                {
                    "eventClassId": 2,
                    "eventId": "1211542",
                    "name": "中国华夏家博会(贵阳)",
                    "countdownDate": 0,
                    "startDate": "2026-06-27",
                    "endDate": "2026-06-28",
                    "address": "贵阳国际会议展览中心",
                    "distance": 9.11,
                }
            ]
        },
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Collect Meituan nearby events.")
    parser.add_argument("--url", default=os.environ.get("MEITUAN_NEARBY_EVENT_URL") or build_url())
    parser.add_argument("--input-json", help="Parse a saved API response instead of requesting.")
    parser.add_argument("--self-test", action="store_true")
    parser.add_argument("--no-db", action="store_true")
    args = parser.parse_args()

    if args.self_test:
        payload = sample_payload()
    elif args.input_json:
        payload = json.loads(Path(args.input_json).read_text(encoding="utf-8-sig"))
    else:
        if not MEITUAN_EB_COOKIE:
            raise RuntimeError("MEITUAN_EB_COOKIE is empty")
        payload = MeituanNearbyEventClient(MEITUAN_EB_COOKIE, args.url).fetch()

    rows = normalize_rows(payload)
    save_outputs(rows, sync_db=not (args.no_db or args.self_test))
    print(f"nearby_event rows={len(rows)}")
    print(f"Excel saved: {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
