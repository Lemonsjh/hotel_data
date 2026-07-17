import argparse
import json
import os
import re
import sys
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from meituan_config import MEITUAN_ME_COOKIE, PARTNER_ID, POI_ID, USER_AGENT

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from ota_mysql_writer import OUTPUT_DIR, sync_table


DEFAULT_OUTPUT = OUTPUT_DIR / "meituan_promotion_capture.json"
DEFAULT_REGISTERED_OUTPUT = OUTPUT_DIR / "meituan_registered_activity_capture.json"
DEFAULT_EXCEL_OUTPUT = OUTPUT_DIR / "meituan_promotion_activity.xlsx"
SPLIT_OUTPUT_DIR = OUTPUT_DIR
DEFAULT_BUSINESS_SCRIPT = Path(__file__).with_name("bussiness_data.py")
COOKIE = MEITUAN_ME_COOKIE
HOTEL_ID = os.environ.get("HOTEL_ID", "").strip()

PROMOTION_TYPE_LABELS = {
    "LAST_MINUTE_SPECIAL": "今夜特价",
    "EVERY_DAY_SPECIAL": "天天特价",
    "EARLY_BOOKING_PREFERENTIAL": "早订多减",
    "FIRST_TIME_CUSTOMER_DISCOUNT": "首住折扣",
    "CONSECUTIVE_STAY_PREFERENTIAL": "连住优惠",
    "TIME_LIMIT_SPECIAL": "限时特惠",
    "FULL_ROOM_ACCELERATOR": "满房加速器",
}

PROMOTION_STATUS_LABELS = {
    "RUNNING": "进行中",
    "PAUSED": "已暂停",
    "PAUSED_TODAY": "跳过当日",
    "CANCEL": "已取消",
    "FINISH": "已结束",
    "AUDIT": "审核中",
    "AUDIT_REVOKE": "撤销退出",
}

PROMOTION_MODE_LABELS = {
    "REDUCE": "立减",
    "DISCOUNT": "折扣",
    "FIX_PRICE": "一口价",
}

STATUS_LISTS = {
    "RUNNING": ["RUNNING"],
    "PAUSED": ["PAUSED", "PAUSED_TODAY"],
    "CANCEL": ["FINISH", "CANCEL"],
}

MODEL_KEY_BY_TYPE = {
    "EARLY_BOOKING_PREFERENTIAL": "preBookModel",
    "CONSECUTIVE_STAY_PREFERENTIAL": "continuesModel",
    "EVERY_DAY_SPECIAL": "dailySpecialModel",
    "LAST_MINUTE_SPECIAL": "tonightSpecialModel",
    "FIRST_TIME_CUSTOMER_DISCOUNT": "firstSpecialModel",
    "TIME_LIMIT_SPECIAL": "timeLimitSpecialModel",
    "FULL_ROOM_ACCELERATOR": "fullRoomModel",
}

WEEKDAY_LABELS = {
    1: "周一",
    2: "周二",
    3: "周三",
    4: "周四",
    5: "周五",
    6: "周六",
    7: "周日",
}

REGISTERED_ACTIVITY_STATUS_LABELS = {
    1: "已参与",
    2: "已结束",
    3: "处理中",
    4: "取消中",
}


class MeituanPromotionClient:
    def __init__(self, cookie: str):
        self.base_url = "https://me.meituan.com"
        self.headers = {
            "User-Agent": USER_AGENT,
            "Accept": "application/json, text/plain, */*",
            "Content-Type": "application/json;charset=UTF-8",
            "Referer": "https://me.meituan.com/ebooking/merchant/promotion/manage",
            "Origin": "https://me.meituan.com",
            "Cookie": cookie,
        }

    def _post(self, path: str, payload: dict[str, Any]) -> Any:
        url = self.base_url + path
        resp = requests.post(url, json=payload, headers=self.headers, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        return unwrap_response(data)

    def _get(self, path: str, params: dict[str, Any]) -> Any:
        url = self.base_url + path
        resp = requests.get(url, params=params, headers=self.headers, timeout=30)
        resp.raise_for_status()
        data = resp.json()
        return unwrap_response(data)

    def query_self_promotion(
        self,
        poi_id: str = POI_ID,
        partner_id: str = PARTNER_ID,
        status: str = "RUNNING",
        promotion_type: str | None = None,
    ) -> Any:
        payload = {
            "partnerId": partner_id,
            "poiId": poi_id,
            "promotionStatusList": STATUS_LISTS.get(status, [status]),
            "promotionActiveTypeEnumList": [promotion_type] if promotion_type else None,
            "queryScene": "MANAGE",
        }
        return self._post("/api/gw/v1/promotion/self/querySelfPromotion?_optSource=ME_PC", payload)

    def query_goods_base_info(
        self,
        poi_id: str = POI_ID,
        partner_id: str = PARTNER_ID,
        status: str = "RUNNING",
        promotion_type: str | None = None,
    ) -> Any:
        payload = {
            "partnerId": partner_id,
            "poiId": poi_id,
            "promotionStatusList": STATUS_LISTS.get(status, [status]),
            "promotionActiveTypeEnumList": [promotion_type] if promotion_type else None,
            "queryScene": "MANAGE",
        }
        return self._post("/api/gw/v1/promotion/self/queryGoodsBaseInfo?_optSource=ME_PC", payload)

    def query_registered_activities(
        self,
        poi_id: str = POI_ID,
        partner_id: str = PARTNER_ID,
        activity_status: int = 1,
        page_num: int = 1,
        page_size: int = 100,
    ) -> Any:
        params = {
            "activityStatus": activity_status,
            "pageSize": page_size,
            "pageNum": page_num,
            "partnerId": partner_id,
            "poiId": poi_id,
            "requestSource": "PC",
        }
        return self._get("/api/gw/activity/getActiveListByStatus", params)

    def query_registered_activity_detail(
        self,
        activity_id: Any,
        poi_id: str = POI_ID,
        partner_id: str = PARTNER_ID,
    ) -> Any:
        params = {
            "activityId": activity_id,
            "partnerId": partner_id,
            "poiId": poi_id,
        }
        return self._get("/api/gw/v1/activity/getActiveDetailByVpoiAndActiveId", params)


def unwrap_response(data: Any) -> Any:
    if not isinstance(data, dict):
        return data
    error = data.get("error") if isinstance(data.get("error"), dict) else {}
    message = (
        data.get("message")
        or data.get("msg")
        or error.get("message")
        or error.get("msg")
        or error.get("displayMsg")
    )
    if data.get("status") not in (None, 0):
        raise RuntimeError({"status": data.get("status"), "message": message})
    if data.get("code") not in (None, 0, 200, 10000, "0", "200", "10000"):
        raise RuntimeError({"code": data.get("code"), "message": message})
    if "data" in data and isinstance(data["data"], (dict, list)):
        return data["data"]
    return data


def money_or_discount(mode: str, value: Any) -> str:
    if value in (None, "", 0):
        return ""
    try:
        number = float(value)
    except (TypeError, ValueError):
        return str(value)
    if mode == "DISCOUNT":
        return f"{number / 10:g}折"
    if mode == "REDUCE":
        return f"减{number / 100:g}元"
    if mode == "FIX_PRICE":
        return f"一口价{number / 100:g}元"
    return str(value)


def time_range(model: dict[str, Any]) -> str:
    if not model:
        return ""
    start = model.get("startTimeStr") or model.get("startTime") or ""
    end = model.get("endTimeStr") or model.get("endTime") or ""
    if start or end:
        return f"{start} 至 {end}".strip()
    return ""


def weekday_text(values: Any) -> str:
    if not isinstance(values, list) or not values:
        return ""
    labels = [WEEKDAY_LABELS.get(int(item), str(item)) for item in values]
    if len(labels) == 7:
        return "整周"
    return "、".join(labels)


def day_hour_text(item: dict[str, Any]) -> str:
    model = item.get("dayHourModel") or {}
    low = str(model.get("low", "")).zfill(4)
    high = str(model.get("high", "")).zfill(4)
    low = f"{low[:2]}:{low[2:]}" if low.strip("0") else ""
    high = f"{high[:2]}:{high[2:]}" if high.strip("0") else ""
    if low or high:
        return f"{low}-{'' if model.get('onlyToday') else '次日'}{high}"
    return ""


def rule_text(promotion_type: str, mode: str, model: dict[str, Any], show_time_discount: bool = False) -> list[str]:
    if not model:
        return []
    prefix = weekday_text(model.get("weekSettingList"))
    rows: list[str] = []
    if promotion_type in ("EVERY_DAY_SPECIAL", "FIRST_TIME_CUSTOMER_DISCOUNT"):
        rows.append(join_rule(prefix, money_or_discount(mode, model.get("promotionData"))))
    elif promotion_type == "FULL_ROOM_ACCELERATOR":
        steps = model.get("stepRuleModelList") or []
        if show_time_discount:
            rows.append(join_rule(prefix, full_room_step_text(steps)))
        else:
            values = sorted({item.get("promotionData") for item in steps if item.get("promotionData")})
            rows.append(join_rule(prefix, "、".join(money_or_discount(mode, item) for item in values)))
    elif promotion_type == "LAST_MINUTE_SPECIAL":
        for item in model.get("stepRuleModelList") or []:
            rows.append(join_rule(prefix, day_hour_text(item), money_or_discount(mode, item.get("promotionData"))))
    elif promotion_type == "TIME_LIMIT_SPECIAL":
        active_time = time_range(model.get("activeTimeModel") or {})
        for item in model.get("stepRuleModelList") or []:
            rows.append(join_rule(active_time, prefix, day_hour_text(item), money_or_discount(mode, item.get("promotionData"))))
    elif promotion_type == "EARLY_BOOKING_PREFERENTIAL":
        parts = [
            f"提前{item.get('preBookDay')}天{money_or_discount(mode, item.get('promotionData'))}"
            for item in model.get("stepRuleModelList") or []
        ]
        rows.append(join_rule(prefix, "、".join(part for part in parts if part)))
    elif promotion_type == "CONSECUTIVE_STAY_PREFERENTIAL":
        parts = [
            f"连住{item.get('continuesDay')}天{money_or_discount(mode, item.get('promotionData'))}"
            for item in model.get("stepRuleModelList") or []
        ]
        rows.append(join_rule(prefix, "、".join(part for part in parts if part)))
    return [row for row in rows if row]


def full_room_step_text(steps: list[dict[str, Any]]) -> str:
    labels = [
        ("promotionData", "周中折扣"),
        ("weekendPromotionData", "周末折扣"),
        ("festivalPromotionData", "节假日折扣"),
    ]
    rows = []
    for key, label in labels:
        values = sorted({item.get(key) for item in steps if item.get(key)})
        if values:
            rows.append(f"{label}：" + "、".join(f"{float(value) / 10:g}折" for value in values))
    return "，".join(rows)


def join_rule(*parts: str) -> str:
    return "，".join(str(part) for part in parts if part)


def normalize_goods(group: dict[str, Any]) -> list[dict[str, Any]]:
    promotion_type = (group.get("promotionModel") or {}).get("promotionActiveType")
    step_goods = {}
    for rule in (group.get("promotionModel") or {}).get("promotionRuleModelList") or []:
        for special in rule.get("specialTimeList") or []:
            model = special.get("fullRoomModel") or {}
            for step in model.get("stepRuleModelList") or []:
                if step.get("goodsId") is not None:
                    step_goods[str(step.get("goodsId"))] = step
    rows = []
    for item in group.get("goodsInfoList") or []:
        ota_product_id = str(item.get("goodsId", ""))
        step = step_goods.get(ota_product_id) or {}
        stock = item.get("saleNum") if promotion_type != "FULL_ROOM_ACCELERATOR" else step.get("stock", item.get("saleNum"))
        rows.append(
            {
                "商品ID": ota_product_id,
                "参与产品": item.get("goodsName") or item.get("name") or item.get("roomName") or item.get("realRoomName") or "",
                "库存": stock if stock not in (None, "", 0) else "无限制",
                "优惠原始值": step.get("promotionData"),
            }
        )
    return rows


def apply_goods_base_info(data: Any, goods_base_info: Any) -> Any:
    if not goods_base_info:
        return data
    goods_map = {}
    if isinstance(goods_base_info, dict):
        candidates = goods_base_info.get("goodsBaseInfoList") or goods_base_info.get("goodsList") or goods_base_info.get("list") or []
    else:
        candidates = goods_base_info if isinstance(goods_base_info, list) else []
    for item in candidates:
        if isinstance(item, dict) and item.get("goodsId") is not None:
            goods_map[str(item.get("goodsId"))] = item.get("goodsName") or item.get("name")
    groups = data.get("goodsGroupInfoList") if isinstance(data, dict) else data
    for group in groups or []:
        for goods in group.get("goodsInfoList") or []:
            name = goods_map.get(str(goods.get("goodsId")))
            if name:
                goods["goodsName"] = name
    return data


def normalize_self_promotions(data: Any) -> list[dict[str, Any]]:
    groups = []
    if isinstance(data, dict):
        groups = data.get("goodsGroupInfoList") or data.get("promotionList") or data.get("list") or []
    elif isinstance(data, list):
        groups = data
    rows = []
    for group in groups:
        model = group.get("promotionModel") or {}
        promotion_type = model.get("promotionActiveType") or group.get("promotionActiveType") or ""
        mode = model.get("promotionType") or ""
        status = model.get("promotionStatus") or group.get("promotionStatus") or ""
        rules = []
        activity_ranges = []
        unable_ranges = []
        for rule in model.get("promotionRuleModelList") or []:
            participate = rule.get("participateTime") or {}
            activity_ranges.append(time_range(participate))
            for item in rule.get("unableTimeList") or []:
                unable_ranges.append(time_range(item))
            for special in rule.get("specialTimeList") or []:
                model_key = MODEL_KEY_BY_TYPE.get(promotion_type, "")
                rules.extend(rule_text(promotion_type, mode, special.get(model_key) or {}, model.get("showTimeDiscount")))
        goods = normalize_goods(group)
        rows.append(
            {
                "促销ID": group.get("promotionId"),
                "促销类型": PROMOTION_TYPE_LABELS.get(promotion_type, promotion_type),
                "促销类型编码": promotion_type,
                "促销状态": PROMOTION_STATUS_LABELS.get(status, status),
                "促销状态编码": status,
                "优惠模式": PROMOTION_MODE_LABELS.get(mode, mode),
                "优惠模式编码": mode,
                "活动时间": "；".join(sorted({item for item in activity_ranges if item})),
                "不可用日期": "；".join(sorted({item for item in unable_ranges if item})),
                "促销规则": "；".join(rules),
                "是否自动延期": "是" if model.get("autoDelay") else "否",
                "是否可编辑": "是" if group.get("canEdit") else "否",
                "是否需要审核": "是" if group.get("needAudit") else "否",
                "参与产品数": len(goods),
                "参与产品": goods,
            }
        )
    return rows


def normalize_registered_goods(goods_list: Any) -> list[dict[str, Any]]:
    rows = []
    for item in goods_list or []:
        if not isinstance(item, dict):
            continue
        rows.append(
            {
                "商品ID": item.get("goodsId") or item.get("productIdStr"),
                "产品名称": item.get("goodsName") or "",
                "房型": item.get("roomType") or "",
                "库存": item.get("stock") or "",
            }
        )
    return rows


def normalize_registered_rule(rule: dict[str, Any]) -> str:
    selected = rule.get("selectedInfo") if isinstance(rule.get("selectedInfo"), dict) else {}
    selected_value = selected.get("value")
    suffix = rule.get("ruleValueSuffix") or ""
    prefix = rule.get("ruleValuePrefix") or ""
    condition = "".join(
        str(part)
        for part in (
            rule.get("ruleConditionPrefix") or "",
            rule.get("ruleCondition") or "",
            rule.get("ruleConditionSuffix") or "",
        )
        if part
    )
    if selected_value not in (None, ""):
        value = f"{prefix}{selected_value}{suffix}".strip()
        return "，".join(part for part in (condition, value) if part)
    values = []
    for item in rule.get("ruleValueList") or []:
        if isinstance(item, dict) and item.get("value") not in (None, ""):
            values.append(f"{item.get('value')}{suffix}")
    if values:
        value = f"{prefix}{'、'.join(values)}".strip()
        return "，".join(part for part in (condition, value) if part)
    return condition


def normalize_registered_detail(detail: Any) -> dict[str, Any]:
    if not isinstance(detail, dict):
        return {}
    rules = [
        normalize_registered_rule(rule)
        for rule in detail.get("promotionRuleList") or []
        if isinstance(rule, dict)
    ]
    return {
        "活动口号": detail.get("activitySlogan") or "",
        "活动说明": detail.get("description") or "",
        "活动类型": detail.get("activityType") or "",
        "折扣说明": detail.get("discountDetail") or "",
        "当前折扣规则": "；".join(rule for rule in rules if rule),
        "叠加规则": detail.get("overlayRule") or "",
        "补充规则": detail.get("supplementaryRule") or "",
        "成本来源": detail.get("costSource") or "",
        "退出规则": detail.get("exitRule") or "",
    }


def normalize_registered_activities(data: Any, details_by_id: dict[Any, Any] | None = None) -> list[dict[str, Any]]:
    details_by_id = details_by_id or {}
    if isinstance(data, dict):
        activities = data.get("activityAppliedInfoList") or data.get("list") or []
    elif isinstance(data, list):
        activities = data
    else:
        activities = []
    rows = []
    for item in activities:
        if not isinstance(item, dict):
            continue
        status = item.get("activityStatus")
        goods = normalize_registered_goods(item.get("appliedGoodsInfoList"))
        exit_info = item.get("exitAuditInfo") if isinstance(item.get("exitAuditInfo"), dict) else {}
        block_info = item.get("activityExtraBlockInfo") if isinstance(item.get("activityExtraBlockInfo"), dict) else {}
        detail = normalize_registered_detail(details_by_id.get(item.get("activityId")))
        rows.append(
            {
                "活动ID": item.get("activityId"),
                "活动名称": item.get("activityName") or "",
                "活动状态": REGISTERED_ACTIVITY_STATUS_LABELS.get(status, str(status) if status is not None else ""),
                "活动状态编码": status,
                "活动类型": detail.get("活动类型") or item.get("activityTypeDesc") or item.get("activityType") or "",
                "活动口号": detail.get("活动口号") or "",
                "活动说明": detail.get("活动说明") or "",
                "折扣说明": detail.get("折扣说明") or "",
                "当前折扣规则": detail.get("当前折扣规则") or "",
                "优惠适用时间": item.get("discountApplicableTime") or "",
                "叠加规则": detail.get("叠加规则") or "",
                "补充规则": detail.get("补充规则") or "",
                "成本来源": detail.get("成本来源") or "",
                "退出规则": detail.get("退出规则") or "",
                "参与产品口径": item.get("participateProduct") or "",
                "是否需要审核": "是" if item.get("isNeedAudit") else "否",
                "是否新数据": "是" if item.get("isNewData") else "否",
                "退出提示": exit_info.get("exitDisplayInfo") or "",
                "是否被新活动阻塞": "是" if block_info.get("needBlockActivity") else "否",
                "替代活动ID": block_info.get("newActiveId"),
                "替代活动名称": block_info.get("newActiveName") or "",
                "参与产品数": len(goods),
                "参与产品": goods,
            }
        )
    return rows


def load_cookie(args: argparse.Namespace) -> str:
    if args.cookie:
        return args.cookie.strip()
    if COOKIE and not COOKIE.startswith("请在这里"):
        return COOKIE.strip()
    if args.cookie_from_business_data:
        text = DEFAULT_BUSINESS_SCRIPT.read_text(encoding="utf-8")
        match = re.search(r'COOKIE\s*=\s*"""(.*?)"""', text, re.S)
        if not match:
            raise RuntimeError("没有在 bussiness_data.py 中找到 COOKIE 变量")
        return match.group(1).strip()
    raise RuntimeError("请先在 meituan_config.py 中配置 MEITUAN_ME_COOKIE 或 MEITUAN_COOKIE。")


def save_json(payload: dict[str, Any], output: Path) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def json_safe(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if hasattr(value, "isoformat") and not isinstance(value, (str, int, float, bool)):
        return value.isoformat()
    return value


def write_standard_json(output_path: Path, headers: list[str], rows: list[list[Any]]) -> Path:
    json_path = Path(output_path).with_suffix(".json")
    payload = {
        "table_name": json_path.stem,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "row_count": len(rows),
        "rows": [
            {header: json_safe(row[index]) if index < len(row) else "" for index, header in enumerate(headers)}
            for row in rows
        ],
    }
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return json_path


def join_unique(values: list[Any]) -> str:
    result = []
    for value in values:
        if value and value not in result:
            result.append(str(value))
    return "、".join(result)


def append_sheet(ws: Any, headers: list[str], rows: list[list[Any]]) -> None:
    ws.append(headers)
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
    for row in rows:
        ws.append(row)
    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for cell in ws[get_column_letter(col_idx)]:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, min(len(value), 50))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(12, max_len + 2)


def build_excel_rows(captured_at: str, self_rows: list[dict[str, Any]], registered_rows: list[dict[str, Any]]) -> tuple[list[list[Any]], list[list[Any]]]:
    summary_rows = []
    detail_rows = []
    for row in self_rows:
        goods = row.get("参与产品") or []
        goods_names = [item.get("参与产品") for item in goods if isinstance(item, dict)]
        summary_rows.append(
            [
                captured_at,
                "美团",
                "自助促销",
                row.get("促销ID"),
                row.get("促销类型"),
                row.get("促销状态"),
                row.get("活动时间"),
                row.get("促销规则"),
                row.get("参与产品数"),
                join_unique(goods_names),
            ]
        )
        for goods_item in goods:
            if not isinstance(goods_item, dict):
                continue
            detail_rows.append(
                [
                    captured_at,
                    "美团",
                    "自助促销",
                    row.get("促销ID"),
                    row.get("促销类型"),
                    goods_item.get("商品ID"),
                    goods_item.get("参与产品"),
                    goods_item.get("库存"),
                ]
            )
    for row in registered_rows:
        goods = row.get("参与产品") or []
        goods_names = [
            item.get("房型") or item.get("产品名称")
            for item in goods
            if isinstance(item, dict)
        ]
        summary_rows.append(
            [
                captured_at,
                "美团",
                "报名活动",
                row.get("活动ID"),
                row.get("活动名称"),
                row.get("活动状态"),
                row.get("优惠适用时间"),
                row.get("当前折扣规则") or row.get("折扣说明"),
                row.get("参与产品数"),
                join_unique(goods_names),
            ]
        )
        for goods_item in goods:
            if not isinstance(goods_item, dict):
                continue
            detail_rows.append(
                [
                    captured_at,
                    "美团",
                    "报名活动",
                    row.get("活动ID"),
                    row.get("活动名称"),
                    goods_item.get("商品ID"),
                    goods_item.get("房型") or goods_item.get("产品名称"),
                    goods_item.get("库存"),
                ]
            )
    return summary_rows, detail_rows


def save_excel(captured_at: str, self_rows: list[dict[str, Any]], registered_rows: list[dict[str, Any]], output: Path = DEFAULT_EXCEL_OUTPUT) -> None:
    output.parent.mkdir(parents=True, exist_ok=True)
    summary_headers = ["snapshot_time", "channel_source", "activity_source_type", "activity_id", "activity_name", "activity_status", "activity_time_range", "activity_rule_labels", "activity_room_type_count", "activity_room_type_summary", "hotel_id"]
    detail_headers = ["snapshot_time", "channel_source", "activity_source_type", "activity_id", "activity_name", "ota_product_id", "room_type_name", "remaining_inventory", "hotel_id"]
    summary_rows, detail_rows = build_excel_rows(captured_at, self_rows, registered_rows)
    summary_rows = [list(row) + [HOTEL_ID] for row in summary_rows]
    detail_rows = [list(row) + [HOTEL_ID] for row in detail_rows]
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "促销活动"
    append_sheet(ws_summary, summary_headers, summary_rows)
    ws_detail = wb.create_sheet("活动产品明细")
    append_sheet(ws_detail, detail_headers, detail_rows)
    wb.save(output)
    summary_path = save_single_sheet(wb, "促销活动", "ota_promotion_activity.xlsx")
    detail_path = save_single_sheet(wb, "活动产品明细", "ota_activity_product_detail.xlsx")
    write_standard_json(summary_path, summary_headers, summary_rows)
    write_standard_json(detail_path, detail_headers, detail_rows)
    sync_table("meituan_ota_promotion_activity", summary_headers, summary_rows, allow_empty_replace=True)
    sync_table("meituan_ota_activity_product_detail", detail_headers, detail_rows, allow_empty_replace=True)


def save_single_sheet(wb: Workbook, sheet_name: str, filename: str) -> Path:
    SPLIT_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    src_ws = wb[sheet_name]
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = sheet_name
    for row in src_ws.iter_rows():
        for cell in row:
            out_cell = out_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                out_cell.font = copy(cell.font)
                out_cell.fill = copy(cell.fill)
                out_cell.border = copy(cell.border)
                out_cell.alignment = copy(cell.alignment)
                out_cell.number_format = cell.number_format
                out_cell.protection = copy(cell.protection)
            if cell.hyperlink:
                out_cell.hyperlink = copy(cell.hyperlink)
            if cell.comment:
                out_cell.comment = copy(cell.comment)
    for key, dimension in src_ws.column_dimensions.items():
        out_ws.column_dimensions[key].width = dimension.width
    for key, dimension in src_ws.row_dimensions.items():
        out_ws.row_dimensions[key].height = dimension.height
    out_ws.freeze_panes = src_ws.freeze_panes
    out_path = SPLIT_OUTPUT_DIR / filename
    out_wb.save(out_path)
    return out_path


def print_summary(rows: list[dict[str, Any]], output: Path) -> None:
    print(f"OK 促销活动数={len(rows)} 输出={output}")
    for index, row in enumerate(rows, 1):
        goods = row.get("参与产品") or []
        goods_names = [item.get("参与产品") for item in goods if item.get("参与产品")]
        print(f"\n[{index}] {row.get('促销类型')} / {row.get('促销状态')}")
        print(f"促销ID：{row.get('促销ID')}")
        print(f"活动时间：{row.get('活动时间') or '-'}")
        print(f"促销规则：{row.get('促销规则') or '-'}")
        print(f"是否自动延期：{row.get('是否自动延期')}")
        print(f"参与产品数：{row.get('参与产品数')}")
        print(f"参与产品：{'、'.join(goods_names) if goods_names else '-'}")


def print_registered_summary(rows: list[dict[str, Any]], output: Path) -> None:
    print(f"OK 已参与报名活动数={len(rows)} 输出={output}")
    for index, row in enumerate(rows, 1):
        goods = row.get("参与产品") or []
        room_names = []
        for item in goods:
            name = item.get("房型") or item.get("产品名称")
            if name and name not in room_names:
                room_names.append(name)
        print(f"\n[{index}] {row.get('活动名称')} / {row.get('活动状态')}")
        print(f"活动ID：{row.get('活动ID')}")
        print(f"活动类型：{row.get('活动类型') or '-'}")
        print(f"折扣说明：{row.get('折扣说明') or '-'}")
        print(f"当前折扣规则：{row.get('当前折扣规则') or '-'}")
        print(f"优惠适用时间：{row.get('优惠适用时间') or '-'}")
        print(f"参与产品口径：{row.get('参与产品口径') or '-'}")
        print(f"参与产品数：{row.get('参与产品数')}")
        print(f"参与房型：{'、'.join(room_names) if room_names else '-'}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Meituan promotion read-only crawler test.")
    parser.add_argument(
        "--tab",
        default="all",
        choices=["all", "self", "registered"],
        help="all=两类都抓；self=已设置的自助促销；registered=已参与的报名活动",
    )
    parser.add_argument("--poi-id", default=POI_ID)
    parser.add_argument("--partner-id", default=PARTNER_ID)
    parser.add_argument("--status", default="RUNNING", choices=["RUNNING", "PAUSED", "CANCEL"])
    parser.add_argument("--activity-status", type=int, default=1, help="已参与报名活动状态，1=已参与，2=已结束")
    parser.add_argument("--promotion-type", default="")
    parser.add_argument("--output", default="")
    parser.add_argument("--excel-output", default=str(DEFAULT_EXCEL_OUTPUT))
    parser.add_argument("--cookie", default="")
    parser.add_argument("--cookie-from-business-data", action="store_true")
    parser.add_argument("--try-goods-base-info", action="store_true", help="尝试调用商品基础信息补全接口；该接口可能返回 403。")
    return parser.parse_args()


def collect_self_promotions(args: argparse.Namespace, client: MeituanPromotionClient) -> list[dict[str, Any]]:
    promotion_raw = client.query_self_promotion(
        poi_id=args.poi_id,
        partner_id=args.partner_id,
        status=args.status,
        promotion_type=args.promotion_type or None,
    )
    goods_base_info = None
    if args.try_goods_base_info:
        try:
            goods_base_info = client.query_goods_base_info(
                poi_id=args.poi_id,
                partner_id=args.partner_id,
                status=args.status,
                promotion_type=args.promotion_type or None,
            )
        except (requests.RequestException, RuntimeError) as exc:
            print(f"WARN 商品基础信息接口不可用，已跳过补全：{exc}")
    promotion_raw = apply_goods_base_info(promotion_raw, goods_base_info)
    rows = normalize_self_promotions(promotion_raw)
    payload = {
        "status": "ok",
        "captured_at": datetime.now().isoformat(timespec="seconds"),
        "platform": "美团",
        "page": "已设置的自助促销",
        "query_status": args.status,
        "promotion_count": len(rows),
        "data": rows,
    }
    output = Path(args.output) if args.output else DEFAULT_OUTPUT
    save_json(payload, output)
    print_summary(rows, output)
    return rows


def collect_registered_activities(args: argparse.Namespace, client: MeituanPromotionClient) -> list[dict[str, Any]]:
    raw = client.query_registered_activities(
        poi_id=args.poi_id,
        partner_id=args.partner_id,
        activity_status=args.activity_status,
        page_num=1,
        page_size=100,
    )
    activities = raw.get("activityAppliedInfoList") if isinstance(raw, dict) else raw
    details_by_id = {}
    for item in activities or []:
        if not isinstance(item, dict) or not item.get("activityId"):
            continue
        try:
            details_by_id[item["activityId"]] = client.query_registered_activity_detail(
                activity_id=item["activityId"],
                poi_id=args.poi_id,
                partner_id=args.partner_id,
            )
        except (requests.RequestException, RuntimeError) as exc:
            print(f"WARN 活动详情接口不可用，已跳过折扣补全 activityId={item.get('activityId')}：{exc}")
    rows = normalize_registered_activities(raw, details_by_id)
    payload = {
        "status": "ok",
        "captured_at": datetime.now().isoformat(timespec="seconds"),
        "platform": "美团",
        "page": "已参与的报名活动",
        "activity_status": args.activity_status,
        "activity_count": len(rows),
        "data": rows,
    }
    output = Path(args.output) if args.output else DEFAULT_REGISTERED_OUTPUT
    save_json(payload, output)
    print_registered_summary(rows, output)
    return rows


def main() -> None:
    args = parse_args()
    cookie = load_cookie(args)
    client = MeituanPromotionClient(cookie)
    captured_at = datetime.now().isoformat(timespec="seconds")
    self_rows: list[dict[str, Any]] = []
    registered_rows: list[dict[str, Any]] = []
    if args.tab == "all":
        self_rows = collect_self_promotions(args, client)
        print("\n" + "=" * 60 + "\n")
        registered_rows = collect_registered_activities(args, client)
    elif args.tab == "registered":
        registered_rows = collect_registered_activities(args, client)
    else:
        self_rows = collect_self_promotions(args, client)
    excel_output = Path(args.excel_output)
    save_excel(captured_at, self_rows, registered_rows, excel_output)
    print(f"\nOK Excel输出={excel_output}")


if __name__ == "__main__":
    main()
