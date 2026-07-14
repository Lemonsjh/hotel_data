from copy import copy

import json
import os
import sys
import requests
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from meituan_config import MEITUAN_EB_COOKIE, PARTNER_ID, POI_ID, USER_AGENT

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from ota_mysql_writer import OUTPUT_DIR, sync_metric_history, sync_table

DEFAULT_EXCEL_PATH = OUTPUT_DIR / "meituan_ota_collected_data.xlsx"
SPLIT_OUTPUT_DIR = OUTPUT_DIR
SHEET_NAME = "采集明细"
HOTEL_ID = os.environ.get("HOTEL_ID", "").strip()
EXCEL_HEADERS = ["snapshot_time", "business_date", "hotel_name", "metric_code", "metric_name",
                 "metric_value", "metric_unit", "compare_label", "compare_value", "competitor_rank", "peer_average"]
METRIC_ORDER = [("PAY_AMT", "销售额"), ("PAY_ORDER_CNT", "支付订单"),
                ("PAY_ROOMNIGHT", "销售间夜"), ("CONSUME_ROOMNIGHT_SPLIT_EX_7DAYS_REFUND", "入住间夜"),
                ("EXPOSE_PV_CNT", "曝光量"), ("INTENTION_UV", "浏览人数"),
                ("PAY_ORDER_CNT_UV", "支付转化率"), ("PAY_ADR", "销售均价"),
                ("DAY_ROOM_LOWEST_PRICE_AVG", "引流价"), ("NOT_AVAILABLE_REAL_ROOM_RATE", "满房率")]
FLOW_METRIC_ORDER = [("FLOW_EXPOSURE_UV", "曝光人数"), ("FLOW_INTENTION_UV", "浏览人数"), ("FLOW_PAY_ORDER_CNT", "支付订单数"),
                     ("FLOW_INTENTION_PER_EXPOSURE", "曝光-浏览转化率"),
                     ("FLOW_PAY_ORDER_PER_INTENTION", "浏览-支付转化率")]
FLOW_FIELD_MAP = {
    "exposureUV": ("FLOW_EXPOSURE_UV", "曝光人数", "人"),
    "intentionUV": ("FLOW_INTENTION_UV", "浏览人数", "人"),
    "payOrderCnt": ("FLOW_PAY_ORDER_CNT", "支付订单数", "单"),
    "intentionPerExposure": ("FLOW_INTENTION_PER_EXPOSURE", "曝光-浏览转化率", ""),
    "payOrderPerIntention": ("FLOW_PAY_ORDER_PER_INTENTION", "浏览-支付转化率", ""),
}
SCORE_METRIC_ORDER = [("SCORE_HOS", "HOS分"), ("SCORE_INFO", "信息分"), ("SCORE_EVALUATION", "评价分")]
SCORE_ID_MAP = {"hos": "SCORE_HOS", "info": "SCORE_INFO", "evaluation": "SCORE_EVALUATION"}
FLOW_TREND_ID_MAP = {
    "1": "FLOW_EXPOSURE_UV", "2": "FLOW_INTENTION_UV", "3": "FLOW_PAY_ORDER_CNT",
    "4": "FLOW_INTENTION_PER_EXPOSURE", "5": "FLOW_PAY_ORDER_PER_INTENTION",
}
BUSINESS_METRIC_IDS = {metric_id for metric_id, _ in METRIC_ORDER + FLOW_METRIC_ORDER}
SCORE_METRIC_IDS = {metric_id for metric_id, _ in SCORE_METRIC_ORDER}
COMPARE_SYMBOLS = {"+": "↑", "-": "↓", "持平": "持平"}
class MeituanDashboardClient:
    def __init__(self, cookie: str):
        self.cookie = cookie
        self.base_url = "https://eb.meituan.com/api/v1/ebooking/home/businessData"
        base = "https://eb.meituan.com/api/shepherdGw/bizDatacenter/hotel/eb"
        self.flow_conversion_url = f"{base}/dataCenter/analyse/flowConversion"
        self.flow_trend_url = f"{base}/dataCenter/analyse/flowTrend"
        self.score_url = f"{base}/dataCenter/home/score"
        self.headers = {
            "User-Agent": USER_AGENT,
            "Accept": "application/json, text/plain, */*",
            "Referer": "https://eb.meituan.com/",
            "Origin": "https://eb.meituan.com",
            "Cookie": cookie
        }
    def _get_api(self, url, params):
        resp = requests.get(url, params=params, headers=self.headers, timeout=30)
        resp.raise_for_status()
        resp.encoding = "utf-8"
        data = resp.json()
        if data["status"] != 0:
            raise Exception(data)
        return data
    def get_dashboard(self, poi_id=POI_ID, partner_id=PARTNER_ID, date_range=1):
        params = {
            "poiId": poi_id,
            "partnerId": partner_id,
            "dateRange": date_range,
            "dataScope": "vpoi",
            "deviceType": "1",
            "yodaReady": "h5",
            "csecplatform": "4",
            "csecversion": "4.2.4"
        }
        return self._parse_dashboard(self._get_api(self.base_url, params))
    def get_report_data(self, poi_id=POI_ID, partner_id=PARTNER_ID, date_range=1):
        dashboard = self.get_dashboard(poi_id=poi_id, partner_id=partner_id, date_range=date_range)
        flow_conversion = self.get_flow_conversion(poi_id=poi_id, partner_id=partner_id, date_range=date_range)
        flow_trend = self.get_flow_trend(poi_id=poi_id, partner_id=partner_id, date_range=date_range)
        score = self.get_score(poi_id=poi_id, partner_id=partner_id)
        dashboard["metrics"].update(flow_conversion.get("metrics") or {})
        for metric_id, trend in (flow_trend.get("metrics") or {}).items():
            if metric_id in dashboard["metrics"]:
                dashboard["metrics"][metric_id]["rank"] = trend.get("rank", "-")
        dashboard["metrics"].update(score.get("metrics") or {})
        dashboard["flow_update_time"] = flow_conversion.get("update_time", "")
        dashboard["score_update_date"] = score.get("update_date", "")
        dashboard["score_hotel_name"] = score.get("hotel_name", "")
        return dashboard
    def get_flow_conversion(self, poi_id=POI_ID, partner_id=PARTNER_ID, date_range=1):
        params = {
            "poiId": poi_id,
            "partnerId": partner_id,
            "dateRange": date_range,
            "dataScope": "vpoi",
            "yodaReady": "h5",
            "csecplatform": "4",
            "csecversion": "4.2.4"
        }
        return self._parse_flow_conversion(self._get_api(self.flow_conversion_url, params))
    def get_flow_trend(self, poi_id=POI_ID, partner_id=PARTNER_ID, date_range=1):
        params = {
            "poiId": poi_id, "partnerId": partner_id, "dateRange": date_range, "dataScope": "vpoi",
            "yodaReady": "h5", "csecplatform": "4", "csecversion": "4.2.4",
        }
        return self._parse_flow_trend(self._get_api(self.flow_trend_url, params))
    def get_score(self, poi_id=POI_ID, partner_id=PARTNER_ID):
        params = {
            "poiId": poi_id,
            "partnerId": partner_id,
            "yodaReady": "h5",
            "csecplatform": "4",
            "csecversion": "4.2.4"
        }
        return self._parse_score(self._get_api(self.score_url, params))
    def get_today_realtime(self, poi_id=POI_ID, partner_id=PARTNER_ID):
        return self.get_report_data(poi_id=poi_id, partner_id=partner_id, date_range=0)
    def get_yesterday(self, poi_id=POI_ID, partner_id=PARTNER_ID):
        return self.get_report_data(poi_id=poi_id, partner_id=partner_id, date_range=1)
    def _parse_dashboard(self, data):
        payload = data.get("data") or {}
        result = {"update_time": payload.get("rtDataUpdateTime", ""), "metrics": {}}
        cards = payload.get("cards") or []
        for card in cards:
            metric = {
                "id": card.get("id", ""),
                "name": card.get("title", ""),
                "value": card.get("value", ""),
                "unit": card.get("suffix") or card.get("unit") or ""
            }
            for ext in card.get("extAttrs") or []:
                ext_name = ext.get("name")
                if ext_name == "同行排名":
                    metric["rank"] = first_value(ext.get("values"))
                elif ext_name == "同行均值":
                    metric["peer_avg"] = first_value(ext.get("values"))
                elif ext_name in ("较上期", "较上周"):
                    metric["compare"] = ext.get("values") or []
                    metric["compare_label"] = ext_name
            if metric["id"]:
                result["metrics"][metric["id"]] = metric
        return result
    def _parse_flow_conversion(self, data):
        payload = data.get("data") or {}
        index_names = payload.get("indexName") or {}
        my_hotel = payload.get("myHotel") or {}
        peer_avg = payload.get("peerAvg") or {}
        result = {"update_time": format_update_time(payload.get("rtDataUpdateTime", "")), "metrics": {}}
        for raw_key, (metric_id, fallback_name, unit) in FLOW_FIELD_MAP.items():
            name = str(index_names.get(raw_key) or fallback_name).replace(" ", "")
            metric = {
                "id": metric_id,
                "name": name,
                "value": first_value(my_hotel.get(raw_key)),
                "unit": unit,
                "peer_avg": first_value(peer_avg.get(raw_key)),
                "source_key": raw_key,
                "source": "flowConversion"
            }
            result["metrics"][metric_id] = metric
        return result
    def _parse_flow_trend(self, data):
        result = {"metrics": {}}
        for card in (data.get("data") or {}).get("cards") or []:
            metric_id = FLOW_TREND_ID_MAP.get(str(card.get("id") or ""))
            if not metric_id:
                continue
            rank = "-"
            for ext in card.get("extAttrs") or []:
                if ext.get("name") == "\u540c\u884c\u6392\u540d":
                    rank = first_value(ext.get("values"))
                    break
            result["metrics"][metric_id] = {"rank": rank}
        return result
    def _parse_score(self, data):
        payload = data.get("data") or {}
        poi_info = payload.get("poiInfo") or {}
        result = {"update_date": poi_info.get("updateDate", ""), "hotel_name": poi_info.get("name", ""), "metrics": {}}
        for card in payload.get("cards") or []:
            metric_id = SCORE_ID_MAP.get(card.get("id", ""))
            if not metric_id:
                continue
            metric = {
                "id": metric_id,
                "name": card.get("title", ""),
                "value": first_value(card.get("value")),
                "unit": "",
                "source_key": card.get("id", ""),
                "source": "score"
            }
            for ext in card.get("extAttrs") or []:
                if ext.get("name") == "同行排名":
                    metric["rank"] = first_value(ext.get("values"))
            result["metrics"][metric_id] = metric
        return result
def first_value(values: Any, default: str = "-") -> str:
    if isinstance(values, list) and values:
        return normalize_value(values[0])
    if values not in (None, ""):
        return normalize_value(values)
    return default
def normalize_value(value: Any) -> str:
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)
def format_update_time(value: Any) -> str:
    if value in (None, ""):
        return ""
    text = str(value)
    if text.isdigit():
        timestamp = int(text)
        if timestamp > 10_000_000_000:
            timestamp = timestamp / 1000
        return "数据更新时间：" + datetime.fromtimestamp(timestamp).strftime(
            "%Y/%m/%d %H:%M"
        )
    return text
def format_compare(metric: dict[str, Any]) -> str:
    values = metric.get("compare")
    if values is None:
        values = metric.get("week_compare")
    if not values:
        return "-"
    if not isinstance(values, list):
        return str(values)
    if len(values) == 1:
        return str(values[0])
    direction = COMPARE_SYMBOLS.get(str(values[0]), str(values[0]))
    return f"{direction}{values[1]}"
def format_metric(metric: dict[str, Any]) -> str:
    value = metric.get("value", "-")
    unit = metric.get("unit", "")
    return f"{value}{unit}"
def format_peer_avg(metric: dict[str, Any]) -> str:
    peer_avg = metric.get("peer_avg")
    if peer_avg in (None, "", "-"):
        return "-"
    unit = metric.get("unit", "")
    return f"{peer_avg}{unit}"
def parse_excel_date(value):
    if not value:
        return ""
    try:
        return datetime.strptime(str(value), "%Y/%m/%d").date()
    except ValueError:
        return value
def excel_metric_value(value, unit=""):
    text = str(value)
    try:
        if text.endswith("%"):
            return float(text.rstrip("%")) / 100
        if unit == "%":
            return float(text) / 100
        number = float(text)
        return int(number) if number.is_integer() else number
    except ValueError:
        return value
def build_metric_rows(data, period_type, date_range, captured_at):
    today = captured_at.date()
    business_date = today if period_type == "today_realtime" else today - timedelta(days=1)
    rows = []
    for metric_id, metric in (data.get("metrics") or {}).items():
        if metric_id not in BUSINESS_METRIC_IDS:
            continue
        rows.append([captured_at, business_date, data.get("score_hotel_name", ""), metric_id, metric.get("name", ""),
                     excel_metric_value(metric.get("value", "-"), metric.get("unit", "")), metric.get("unit", ""),
                     metric.get("compare_label", ""), format_compare(metric), metric.get("rank", "-"),
                     metric.get("peer_avg", "-")])
    return rows
def build_score_rows(data, captured_at):
    rows = []
    business_date = parse_excel_date(data.get("score_update_date", "")) or captured_at.date() - timedelta(days=1)
    for metric_id in SCORE_METRIC_IDS:
        metric = (data.get("metrics") or {}).get(metric_id)
        if not metric:
            continue
        rows.append([captured_at, business_date, data.get("score_hotel_name", ""), metric_id, metric.get("name", ""),
                     excel_metric_value(metric.get("value", ""), metric.get("unit", "")), metric.get("unit", ""),
                     "", "", metric.get("rank", "-"), "-"])
    return rows
def save_to_excel(today_data, yesterday_data, output_path=DEFAULT_EXCEL_PATH):
    captured_at = datetime.now()
    rows = build_metric_rows(today_data, "today_realtime", 0, captured_at)
    rows += build_metric_rows(yesterday_data, "yesterday", 1, captured_at)
    rows += build_score_rows(today_data, captured_at)
    write_rows_to_workbook(rows, output_path)
    return output_path
def write_rows_to_workbook(rows, output_path):
    headers = [*EXCEL_HEADERS, "hotel_id"]
    rows = [list(row) + [HOTEL_ID] for row in rows]
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(output_path) if output_path.exists() else Workbook()
    if SHEET_NAME in wb.sheetnames: del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME, 0)
    ws.append(headers)
    ws.freeze_panes = "A2"
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")
    for idx, width in enumerate([20, 14, 24, 32, 22, 14, 8, 14, 14, 12, 14, 16], 1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width
    for row in rows:
        ws.append(row)
        row_idx = ws.max_row
        ws.cell(row_idx, 1).number_format = "yyyy-mm-dd hh:mm:ss"
        ws.cell(row_idx, 3).number_format = "yyyy-mm-dd"
        if ws.cell(row_idx, 7).value == "%" or ws.cell(row_idx, 5).value in ("曝光-浏览转化率", "浏览-支付转化率"):
            ws.cell(row_idx, 6).number_format = "0.00%"
    wb.save(output_path)
    save_single_sheet(wb, SHEET_NAME, "ota_business_metrics.xlsx")


def json_safe(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if hasattr(value, "isoformat") and not isinstance(value, (str, int, float, bool)):
        return value.isoformat()
    return value


def write_standard_json(output_path, headers, rows):
    json_path = Path(output_path).with_suffix(".json")
    payload = {
        "table_name": json_path.stem,
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "row_count": len(rows),
        "rows": [
            {header: json_safe(row[index]) if index < len(row) else "" for index, header in enumerate(headers)}
            for row in rows
        ],
    }
    json_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return json_path


def save_single_sheet(wb, sheet_name, filename):
    SPLIT_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    src_ws = wb[sheet_name]
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = sheet_name
    for row in src_ws.iter_rows():
        for cell in row:
            out_cell = out_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                out_cell.font = copy(cell.font)
                out_cell.fill = copy(cell.fill)
                out_cell.border = copy(cell.border)
                out_cell.alignment = copy(cell.alignment)
                out_cell.number_format = cell.number_format
                out_cell.protection = copy(cell.protection)
            if cell.hyperlink:
                out_cell.hyperlink = copy(cell.hyperlink)
            if cell.comment:
                out_cell.comment = copy(cell.comment)
    for key, dimension in src_ws.column_dimensions.items():
        out_ws.column_dimensions[key].width = dimension.width
    for key, dimension in src_ws.row_dimensions.items():
        out_ws.row_dimensions[key].height = dimension.height
    out_ws.freeze_panes = src_ws.freeze_panes
    out_path = SPLIT_OUTPUT_DIR / filename
    out_wb.save(out_path)
    headers = [cell.value for cell in out_ws[1]]
    rows = [list(row) for row in out_ws.iter_rows(min_row=2, values_only=True)]
    write_standard_json(out_path, headers, rows)
    table_name = f"meituan_{Path(filename).stem}"
    if table_name == "meituan_ota_business_metrics":
        sync_metric_history(headers, rows)
    else:
        sync_table(table_name, headers, rows)
    return out_path


def metric_block(metrics: dict[str, dict[str, Any]], metric_id: str, fallback_name: str) -> str:
    metric = metrics.get(metric_id)
    compare_label = "较上周"
    if not metric:
        return f"{fallback_name}：-\\n{compare_label}：-\\n同行排名：-\\n同行均值：-"
    name, compare_label = metric.get("name") or fallback_name, metric.get("compare_label") or compare_label
    return "\n".join([f"{name}：{format_metric(metric)}", f"{compare_label}：{format_compare(metric)}",
             f"同行排名：{metric.get('rank', '-')}", f"同行均值：{format_peer_avg(metric)}"]
    )
def flow_metric_block(metrics: dict[str, dict[str, Any]], metric_id: str, fallback_name: str) -> str:
    metric = metrics.get(metric_id)
    if not metric: return f"{fallback_name}：-\n同行均值：-"
    name = metric.get("name") or fallback_name
    return "\n".join([f"{name}：{format_metric(metric)}", f"同行均值：{format_peer_avg(metric)}"])
def score_metric_block(metrics: dict[str, dict[str, Any]], metric_id: str, fallback_name: str) -> str:
    metric = metrics.get(metric_id)
    if not metric: return f"{fallback_name}：-\n同行排名：-"
    return "\n".join([f"{metric.get('name') or fallback_name}：{format_metric(metric)}",
                      f"同行排名：{metric.get('rank', '-')}"])
def generate_report(data, title):
    metrics = data.get("metrics") or {}
    metric_sections = "\n\n".join(metric_block(metrics, metric_id, name) for metric_id, name in METRIC_ORDER)
    flow_sections = "\n\n".join(flow_metric_block(metrics, metric_id, name) for metric_id, name in FLOW_METRIC_ORDER)
    score_sections = "\n\n".join(score_metric_block(metrics, metric_id, name) for metric_id, name in SCORE_METRIC_ORDER)
    update_lines = [f"更新时间：\n{data.get('update_time', '-')}"]
    if data.get("flow_update_time"):
        update_lines.append(f"流量转化更新时间：\n{data.get('flow_update_time')}")
    if data.get("score_update_date"):
        update_lines.append(f"评分更新日期：\n{data.get('score_update_date')}")
    return "\n".join(["", "========================", title, "========================", "",
                      "\n".join(update_lines), "", metric_sections, "", "------------------------",
                      "流量转化分析", "------------------------", "", flow_sections, "",
                      "------------------------", "最新评分数据", "------------------------", "",
                      score_sections, "", "========================"])
def generate_daily_report(data):
    return generate_report(data, "美团经营日报")
if __name__ == "__main__":
    if not MEITUAN_EB_COOKIE:
        raise RuntimeError("??? meituan_config.py ??? MEITUAN_EB_COOKIE ? MEITUAN_COOKIE?")
    client = MeituanDashboardClient(MEITUAN_EB_COOKIE)
    today_data = client.get_today_realtime()
    yesterday_data = client.get_yesterday()
    print(today_data)
    print(generate_report(today_data, "美团今日实时经营数据"))
    print(yesterday_data)
    print(generate_report(yesterday_data, "美团昨日经营日报"))
    excel_path = save_to_excel(today_data, yesterday_data)
    print(f"Excel已保存：{excel_path}")
